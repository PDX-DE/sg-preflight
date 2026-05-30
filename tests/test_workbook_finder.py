"""H-27 tests for workbook_finder + workbook_generator + delivery_checklist wiring."""
from __future__ import annotations

import json
import os
import tempfile
import time
import unittest
from datetime import datetime, timezone
from pathlib import Path
from unittest import mock

from sg_preflight.workbook_finder import (
    SOURCE_AUTO_GENERATED_LOCALLY,
    SOURCE_FROM_CI,
    WORKBOOK_FORMAT_A_DATE_STAMPED,
    WORKBOOK_FORMAT_B_VERSION_TAGGED,
    WORKBOOK_FORMAT_UNKNOWN,
    _build_search_locations,
    _candidate_profile_ids,
    _format_hint,
    find_workbook_candidates,
    render_resolution_text,
    resolve_workbook,
)


class WorkbookFinderHelperTests(unittest.TestCase):
    def test_candidate_profile_ids_expands_evo_suffix(self) -> None:
        self.assertEqual(_candidate_profile_ids("G70"), ("G70", "G70_EVO"))
        self.assertEqual(_candidate_profile_ids("G70_EVO"), ("G70_EVO", "G70"))
        # Already-lowercase / mixed-case normalize to upper.
        self.assertEqual(_candidate_profile_ids("f70"), ("F70", "F70_EVO"))
        self.assertEqual(_candidate_profile_ids(""), ())

    def test_format_hint_detects_format_a_and_b_filenames(self) -> None:
        self.assertEqual(_format_hint(Path("G65_20251002.xlsx")), WORKBOOK_FORMAT_A_DATE_STAMPED)
        self.assertEqual(_format_hint(Path("F70_vx.xlsx")), WORKBOOK_FORMAT_B_VERSION_TAGGED)
        self.assertEqual(_format_hint(Path("U10_v9.xlsx")), WORKBOOK_FORMAT_B_VERSION_TAGGED)
        # Auto-generated workbooks always emit Format A shape per H-27 spec.
        self.assertEqual(_format_hint(Path("G70_auto_20260529.xlsx")), WORKBOOK_FORMAT_A_DATE_STAMPED)
        # Unknown filenames fall through.
        self.assertEqual(_format_hint(Path("random_workbook.xlsx")), WORKBOOK_FORMAT_UNKNOWN)

    def test_search_locations_cover_eight_documented_paths_per_directive(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            ws = Path(tmp) / "workspace"
            br = Path(tmp) / "bmw"
            locations = _build_search_locations("G70", workspace=ws, bmw_root=br)
            source_keys = {key for key, _d, _g in locations}
            # The eight directive-named slots (auto-gen output dir is bonus operator-local).
            for key in (
                "svn_size_analysis_date_stamped",
                "svn_size_analysis_version_tagged",
                "svn_idcevo_size_analysis",
                "svn_cars_bmw_size_analysis",
                "svn_idc23_export_size_analysis",
                "bmw_git_evo_size_analysis",
                "bmw_git_export_size_analysis",
                "bmw_git_evo_size_analysis_dash",
            ):
                self.assertIn(key, source_keys, f"H-27 directive slot {key!r} missing from search locations")
            # operator_local_auto_gen is bonus (search the auto-gen output dir on re-runs).
            self.assertIn("operator_local_auto_gen", source_keys)

    def test_build_search_locations_does_not_double_evo_suffix(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            ws = Path(tmp) / "workspace"
            br = Path(tmp) / "bmw"
            locations = _build_search_locations("G70_EVO", workspace=ws, bmw_root=br)
            for _key, directory, _glob in locations:
                # `G70_EVO_EVO` is meaningless and should never appear in the search.
                self.assertNotIn("G70_EVO_EVO", str(directory))


class WorkbookFinderResolutionTests(unittest.TestCase):
    def test_resolve_returns_unavailable_when_no_workbook_exists(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            ws = Path(tmp) / "workspace"
            ws.mkdir(parents=True)
            resolution = resolve_workbook("G70", workspace=ws, bmw_root=ws / "bmw")
            self.assertEqual(resolution.status, "unavailable")
            self.assertEqual(resolution.candidates, ())
            self.assertIsNone(resolution.selected)
            # Search paths should still be reported so the operator can verify the
            # finder walked the expected locations.
            self.assertGreaterEqual(len(resolution.search_paths), 8)

    def test_resolve_picks_newest_mtime_across_locations(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            ws = Path(tmp) / "workspace"
            br = Path(tmp) / "bmw"
            # Old workbook in slot 1 (SVN date-stamped).
            old = ws / "Cars" / "size_analysis" / "G70_20250101.xlsx"
            old.parent.mkdir(parents=True, exist_ok=True)
            old.write_bytes(b"old workbook bytes")
            # Backdate the old file so the newer one wins regardless of FS resolution.
            old_ts = time.time() - 3600
            os.utime(old, (old_ts, old_ts))
            # New workbook in slot 5 (BMW Git EVO).
            new = br / "cars" / "BMW" / "G70_EVO" / "export" / "size_analysis" / "G70_20260101.xlsx"
            new.parent.mkdir(parents=True, exist_ok=True)
            new.write_bytes(b"new workbook bytes")
            new_ts = time.time()
            os.utime(new, (new_ts, new_ts))

            resolution = resolve_workbook("G70", workspace=ws, bmw_root=br)
            self.assertEqual(resolution.status, "available")
            self.assertIsNotNone(resolution.selected)
            self.assertEqual(resolution.selected.path.name, "G70_20260101.xlsx")
            self.assertEqual(resolution.selected.source_classification, SOURCE_FROM_CI)
            self.assertEqual(resolution.selected.workbook_format, WORKBOOK_FORMAT_A_DATE_STAMPED)
            # Both workbooks are surfaced as candidates so the operator sees both.
            self.assertEqual(len(resolution.candidates), 2)

    def test_resolve_recognises_auto_generated_classification(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            ws = Path(tmp) / "workspace"
            # Drop a synthetic auto-generated workbook into the documented operator-
            # local output dir. Use mock.patch on Path.home so the test is hermetic.
            home = Path(tmp) / "home"
            auto_dir = home / "sgfx_outputs" / "g70" / "delivery-workbook"
            auto_dir.mkdir(parents=True, exist_ok=True)
            auto_path = auto_dir / "G70_auto_20260529.xlsx"
            auto_path.write_bytes(b"auto generated bytes")
            with mock.patch("sg_preflight.workbook_finder.Path.home", return_value=home):
                resolution = resolve_workbook("G70", workspace=ws, bmw_root=ws / "bmw")
            self.assertEqual(resolution.status, "available")
            self.assertEqual(resolution.selected.source_classification, SOURCE_AUTO_GENERATED_LOCALLY)
            self.assertEqual(resolution.selected.source_key, "operator_local_auto_gen")

    def test_render_resolution_text_emits_operator_readable_lines(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            ws = Path(tmp) / "workspace"
            ws.mkdir()
            unavailable = resolve_workbook("Z99", workspace=ws, bmw_root=ws / "bmw")
            text = render_resolution_text(unavailable)
            self.assertIn("status:       unavailable", text)
            self.assertIn("Z99", text)


class WorkbookGeneratorTests(unittest.TestCase):
    def test_generate_from_raw_emits_format_a_overview_and_provenance_sheet(self) -> None:
        from sg_preflight.workbook_generator import (
            FORMAT_A_HEADER,
            RawExportSizeData,
            RawVariantRow,
            VALEO_COMPRESSION_FACTOR,
            generate_workbook_from_raw,
        )
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmp:
            target = Path(tmp) / "G65_auto_20260529.xlsx"
            data = RawExportSizeData(
                profile_id="G65",
                source_path=Path(tmp) / "raw.json",
                rows=(
                    RawVariantRow("BEV-Basis", 5616, 11935.61, 10677.94, 435.56),
                    RawVariantRow("ICE-Basis", 4200, 9800.0, 8600.0, 380.0),
                ),
            )
            generate_workbook_from_raw(data, output_path=target, today=datetime(2026, 5, 29, tzinfo=timezone.utc))
            self.assertTrue(target.is_file())
            wb = load_workbook(target, read_only=True)
            try:
                self.assertIn("Overview", wb.sheetnames)
                self.assertIn("SGFX Provenance", wb.sheetnames)
                overview_rows = list(wb["Overview"].iter_rows(values_only=True))
                # Row 1: profile id.
                self.assertEqual(overview_rows[0][0], "G65")
                # Row 2: canonical header.
                self.assertEqual(tuple(overview_rows[1]), FORMAT_A_HEADER)
                # Row 3: BEV-Basis with computed Total + Valeo est.
                bev = overview_rows[2]
                self.assertEqual(bev[0], "BEV-Basis")
                self.assertAlmostEqual(float(bev[5]), 28665.11, places=2)
                self.assertAlmostEqual(
                    float(bev[6]),
                    round(28665.11 * VALEO_COMPRESSION_FACTOR, 2),
                    places=2,
                )
                # Provenance sheet carries honest classification.
                prov_rows = {row[0]: row[1] for row in wb["SGFX Provenance"].iter_rows(values_only=True)}
                self.assertEqual(prov_rows.get("source_classification"), SOURCE_AUTO_GENERATED_LOCALLY)
                self.assertEqual(prov_rows.get("workbook_format"), WORKBOOK_FORMAT_A_DATE_STAMPED)
                self.assertEqual(prov_rows.get("variant_count"), 2)
            finally:
                wb.close()

    def test_auto_generate_picks_up_json_raw_data_and_classifies_as_locally_generated(self) -> None:
        from sg_preflight.workbook_generator import auto_generate_if_raw_available

        with tempfile.TemporaryDirectory() as tmp:
            bmw_root = Path(tmp) / "bmw"
            raw_dir = bmw_root / "cars" / "BMW" / "G70_EVO" / "export"
            raw_dir.mkdir(parents=True)
            (raw_dir / "size_data.json").write_text(
                json.dumps({
                    "variants": [
                        {"variant": "BEV-Basis", "TextureCube": 5616, "Texture2D": 11935.61, "ArrayResource": 10677.94, "Effect": 435.56},
                    ]
                }),
                encoding="utf-8",
            )
            with mock.patch("sg_preflight.workbook_generator.Path.home", return_value=Path(tmp) / "home"):
                candidate = auto_generate_if_raw_available("G70", workspace=tmp, bmw_root=bmw_root)
            self.assertIsNotNone(candidate)
            self.assertEqual(candidate.source_classification, SOURCE_AUTO_GENERATED_LOCALLY)
            self.assertEqual(candidate.workbook_format, WORKBOOK_FORMAT_A_DATE_STAMPED)
            self.assertTrue(candidate.path.is_file())
            self.assertIn("delivery-workbook", str(candidate.path))
            self.assertIn("g70", str(candidate.path).lower())

    def test_auto_generate_returns_none_when_no_raw_data(self) -> None:
        from sg_preflight.workbook_generator import auto_generate_if_raw_available

        with tempfile.TemporaryDirectory() as tmp:
            with mock.patch("sg_preflight.workbook_generator.Path.home", return_value=Path(tmp) / "home"):
                candidate = auto_generate_if_raw_available(
                    "G70",
                    workspace=Path(tmp) / "workspace",
                    bmw_root=Path(tmp) / "bmw",
                )
            self.assertIsNone(candidate)

    def test_csv_raw_data_path_also_parses_with_alias_headers(self) -> None:
        from sg_preflight.workbook_generator import find_raw_export_size_data

        with tempfile.TemporaryDirectory() as tmp:
            bmw_root = Path(tmp) / "bmw"
            raw_dir = bmw_root / "cars" / "BMW" / "G70_EVO" / "export"
            raw_dir.mkdir(parents=True)
            (raw_dir / "size_data.csv").write_text(
                "Variant,TextureCube,Texture2D,ArrayResource,Effect\n"
                "BEV-Basis,5616,11935.61,10677.94,435.56\n"
                "ICE-Basis,4200,9800,8600,380\n",
                encoding="utf-8",
            )
            raw = find_raw_export_size_data("G70", workspace=tmp, bmw_root=bmw_root)
            self.assertIsNotNone(raw)
            self.assertEqual(len(raw.rows), 2)
            self.assertEqual(raw.rows[0].variant, "BEV-Basis")
            self.assertAlmostEqual(raw.rows[0].total, 28665.11, places=2)


class DeliveryChecklistWiringTests(unittest.TestCase):
    def test_delivery_checklist_falls_through_to_finder_when_legacy_lookup_misses(self) -> None:
        """H-27 wiring: when the legacy single-path lookup returns None, the new
        multi-location finder must take over and pick up a workbook in any of the
        eight documented slots."""
        from sg_preflight.delivery_checklist import resolve_delivery_checklist_workbook

        with tempfile.TemporaryDirectory() as tmp:
            ws = Path(tmp) / "workspace"
            br = Path(tmp) / "bmw"
            # Drop a workbook ONLY in a slot the legacy lookup ignores
            # (`<bmw_root>/cars/BMW/G70_EVO/export/size_analysis/`).
            target = br / "cars" / "BMW" / "G70_EVO" / "export" / "size_analysis" / "G70_20260529.xlsx"
            target.parent.mkdir(parents=True, exist_ok=True)
            target.write_bytes(b"workbook bytes")

            resolved = resolve_delivery_checklist_workbook(
                workspace=ws,
                profile_id="G70",
                bmw_root=br,
            )
            self.assertEqual(resolved, target.resolve())


if __name__ == "__main__":
    unittest.main()
