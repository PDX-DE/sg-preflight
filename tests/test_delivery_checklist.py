from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from sg_preflight.daily_digest import build_daily_digest
from sg_preflight.delivery_checklist import (
    READ_ONLY_BANNER,
    read_delivery_checklist,
    render_delivery_checklist_markdown,
    resolve_delivery_checklist_workbook,
)


def _write_delivery_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Delivery"
    worksheet.append(
        [
            "Car",
            "Last Tested",
            "SVN Revision",
            "Changelog Revision",
            "Export Size",
            "Screenshots",
            "Interface",
            "Perspectives",
        ]
    )
    worksheet.append(["G65_EVO", "2026-05-14 08:30", "r12345", "r12340", "OK", "Fail", "n/a", "Blocked"])
    worksheet.append(["G70_EVO", "2026-05-13 17:00", "r12330", "r12310", "OK", "OK", "OK", "OK"])
    workbook.save(path)


def _write_export_size_workbook(
    path: Path,
    *,
    latest_date: str = "15.05.2026",
    latest_svn: str = "r12399",
    latest_ramses_size: str = "10.4MB",
    latest_logic_size: str = "1.5MB",
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "G65"
    worksheet.append(["Date", "SVN", "Changelog", "Ramses Size", "Logic Size", "Screenshot", "Comment"])
    worksheet.append(["14.05.2026", "r12345", "r12340", "10.2MB", "1.4MB", "", "fixture row"])
    worksheet.append([latest_date, latest_svn, "r12380", latest_ramses_size, latest_logic_size, "", "latest row"])
    workbook.save(path)


def _write_size_analysis_overview_workbook(
    path: Path,
    *,
    profile_id: str = "G65",
    variant_count: int = 2,
    total_base: float = 28000.0,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Overview"
    worksheet.append([profile_id, "", "", "", "", "", ""])
    worksheet.append(["Variant", "TextureCube", "Texture2D", "ArrayResource", "Effect", "Total", "Valeo est."])
    for index in range(variant_count):
        worksheet.append(
            [
                f"Variant-{index + 1}",
                "5616",
                "11935.61",
                "10677.94",
                "435.56",
                str(total_base + index),
                "25225.3",
            ]
        )
    workbook.save(path)


def _write_versioned_size_analysis_overview_workbook(
    path: Path,
    *,
    profile_id: str = "F66",
    version: str = "vx",
    variant_count: int = 3,
    total_base: float = 12000.0,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Overview"
    worksheet.append([f"{profile_id} Sizes - {version}", "", "Force recalculate with Ctrl-Alt-F9", ""])
    worksheet.append(["", "", "", ""])
    worksheet.append(["", *[f"Variant-{index + 1}" for index in range(variant_count)], "", "Min", "Max"])
    worksheet.append(["Date", *["28.08.24 15.11" for _ in range(variant_count)], "", "", ""])
    worksheet.append(["Total", *[str(total_base + index) for index in range(variant_count)], "", str(total_base), str(total_base + variant_count - 1)])
    worksheet.append(["Textures", *["9000" for _ in range(variant_count)], "", "9000", "9000"])
    worksheet.append(["Meshes", *["3000" for _ in range(variant_count)], "", "3000", "3000"])
    workbook.save(path)


class TestDeliveryChecklist(unittest.TestCase):
    def test_read_delivery_checklist_parses_profile_row_without_writing(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "Delivery Data - BMW.xlsx"
            _write_delivery_workbook(workbook_path)
            before_mtime = workbook_path.stat().st_mtime_ns

            payload = read_delivery_checklist(
                profile_id="G65",
                workbook_path=workbook_path,
                workspace=Path(temp_dir),
            )

            after_mtime = workbook_path.stat().st_mtime_ns

        self.assertEqual(payload["status"], "available")
        self.assertTrue(payload["data_available"])
        self.assertEqual(payload["profile_id"], "G65")
        self.assertEqual(payload["matched_profile_id"], "G65_EVO")
        self.assertEqual(payload["worksheet"], "Delivery")
        self.assertEqual(payload["last_tested"], "2026-05-14 08:30")
        self.assertEqual(payload["svn_revision"], "r12345")
        self.assertEqual(payload["changelog_revision"], "r12340")
        self.assertEqual(payload["workbook_metadata"]["brand"], "BMW")
        self.assertGreaterEqual(payload["workbook_metadata"]["row_count"], 2)
        self.assertGreater(payload["workbook_metadata"]["file_size"], 0)
        self.assertTrue(payload["workbook_metadata"]["modified_at"])
        self.assertFalse(payload["is_approval"])
        self.assertEqual(before_mtime, after_mtime)
        checks = {item["key"]: item for item in payload["checks"]}
        self.assertEqual(checks["export_size"]["status"], "passed")
        self.assertEqual(checks["screenshots"]["status"], "failed")
        self.assertEqual(checks["interface"]["status"], "not_applicable")
        self.assertEqual(checks["perspectives"]["status"], "blocked")
        self.assertIn("read-only", payload["note"].lower())

    def test_read_delivery_checklist_reports_missing_workbook_cleanly(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)

            payload = read_delivery_checklist(profile_id="G65", workspace=root)

        self.assertEqual(payload["status"], "unavailable")
        self.assertFalse(payload["data_available"])
        self.assertEqual(payload["profile_id"], "G65")
        self.assertIn("BMW Export Size.xlsx", payload["workbook_path"])
        self.assertIn("delivery-checklist data unavailable", payload["summary"].lower())
        self.assertEqual(payload["checks"], [])

    def test_missing_workbook_uses_existing_trunk_delivery_checklist_folder(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            trunk = Path(temp_dir) / "repositories" / "trunk"
            checklist_dir = trunk / ".pdx" / "checkers" / "deliveryChecklist"
            checklist_dir.mkdir(parents=True)

            resolved = resolve_delivery_checklist_workbook(workspace=trunk)
            payload = read_delivery_checklist(profile_id="G65", workspace=trunk)

        expected = checklist_dir / "Delivery Data - BMW.xlsx"
        self.assertEqual(resolved, expected.resolve())
        self.assertEqual(Path(payload["workbook_path"]), expected.resolve())
        self.assertNotIn("repositories\\trunk\\repositories\\trunk", payload["workbook_path"])

    def test_read_delivery_checklist_supports_export_size_workbook_shape(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            trunk = Path(temp_dir) / "repositories" / "trunk"
            workbook_path = trunk / "Cars" / "BMW" / "BMW Export Size.xlsx"
            _write_export_size_workbook(workbook_path)

            payload = read_delivery_checklist(profile_id="G65", workspace=trunk)

        self.assertEqual(payload["status"], "available")
        self.assertEqual(payload["matched_profile_id"], "G65")
        self.assertEqual(payload["worksheet"], "G65")
        self.assertEqual(payload["last_tested"], "15.05.2026")
        self.assertEqual(payload["svn_revision"], "r12399")
        self.assertEqual(payload["changelog_revision"], "r12380")
        checks = {item["key"]: item for item in payload["checks"]}
        self.assertEqual(checks["ramses_size"]["status"], "recorded")
        self.assertEqual(checks["ramses_size"]["raw_value"], "10.4MB")
        self.assertEqual(checks["logic_size"]["status"], "recorded")
        self.assertEqual(checks["logic_size"]["raw_value"], "1.5MB")
        self.assertIn("Ramses Size recorded", payload["summary"])
        self.assertFalse(payload["is_approval"])

    def test_read_delivery_checklist_discovers_latest_size_analysis_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            trunk = Path(temp_dir) / "repositories" / "trunk"
            old_workbook = trunk / "Cars" / "size_analysis" / "G65_20250930.xlsx"
            new_workbook = trunk / "Cars" / "size_analysis" / "G65_20251002.xlsx"
            _write_size_analysis_overview_workbook(old_workbook, variant_count=1, total_base=27000.0)
            _write_size_analysis_overview_workbook(new_workbook, variant_count=3, total_base=28000.0)

            resolved = resolve_delivery_checklist_workbook(workspace=trunk, profile_id="G65")
            payload = read_delivery_checklist(profile_id="G65", workspace=trunk)

        self.assertEqual(resolved, new_workbook.resolve())
        self.assertEqual(Path(payload["workbook_path"]), new_workbook.resolve())
        self.assertEqual(payload["status"], "available")
        self.assertEqual(payload["matched_profile_id"], "G65")
        self.assertEqual(payload["worksheet"], "Overview")
        self.assertEqual(payload["last_tested"], "2025-10-02")
        checks = {item["key"]: item for item in payload["checks"]}
        self.assertEqual(checks["variant_count"]["status"], "recorded")
        self.assertEqual(checks["variant_count"]["raw_value"], "3")
        self.assertIn("28000.0", checks["variant_totals"]["raw_value"])
        self.assertIn("3 variant rows", payload["summary"])
        self.assertFalse(payload["is_approval"])

    def test_read_delivery_checklist_discovers_versioned_size_analysis_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            trunk = Path(temp_dir) / "repositories" / "trunk"
            old_workbook = trunk / "Cars" / "size_analysis" / "F66_v9.xlsx"
            new_workbook = trunk / "Cars" / "size_analysis" / "F66_vx.xlsx"
            _write_versioned_size_analysis_overview_workbook(
                old_workbook,
                profile_id="F66",
                version="v9",
                variant_count=1,
                total_base=11000.0,
            )
            _write_versioned_size_analysis_overview_workbook(
                new_workbook,
                profile_id="F66",
                version="vx",
                variant_count=4,
                total_base=12000.0,
            )

            resolved = resolve_delivery_checklist_workbook(workspace=trunk, profile_id="F66")
            payload = read_delivery_checklist(profile_id="F66", workspace=trunk)

        self.assertEqual(resolved, new_workbook.resolve())
        self.assertEqual(Path(payload["workbook_path"]), new_workbook.resolve())
        self.assertEqual(payload["status"], "available")
        self.assertEqual(payload["matched_profile_id"], "F66 Sizes - vx")
        self.assertEqual(payload["worksheet"], "Overview")
        self.assertEqual(payload["last_tested"], "28.08.24 15.11")
        checks = {item["key"]: item for item in payload["checks"]}
        self.assertEqual(checks["variant_count"]["raw_value"], "4")
        self.assertIn("Variant-1=12000.0", checks["variant_totals"]["raw_value"])
        self.assertIn("4 variant columns", payload["summary"])
        self.assertFalse(payload["is_approval"])

    def test_read_delivery_checklist_discovers_profile_common_fallback(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            trunk = Path(temp_dir) / "repositories" / "trunk"
            workbook_path = trunk / "Cars_IDCevo" / "BMW" / "G65" / "_Common" / "Delivery Data - BMW.xlsx"
            _write_delivery_workbook(workbook_path)

            resolved = resolve_delivery_checklist_workbook(workspace=trunk, profile_id="G65")
            payload = read_delivery_checklist(profile_id="G65", workspace=trunk)

        self.assertEqual(resolved, workbook_path.resolve())
        self.assertEqual(Path(payload["workbook_path"]), workbook_path.resolve())
        self.assertEqual(payload["status"], "available")
        self.assertEqual(payload["matched_profile_id"], "G65_EVO")
        self.assertFalse(payload["is_approval"])

    def test_export_size_workbook_not_run_values_are_not_treated_as_recorded(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            trunk = Path(temp_dir) / "repositories" / "trunk"
            workbook_path = trunk / "Cars" / "BMW" / "BMW Export Size.xlsx"
            workbook_path.parent.mkdir(parents=True, exist_ok=True)
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "G65"
            worksheet.append(["Date", "SVN", "Changelog", "Ramses Size", "Logic Size", "Screenshot", "Comment"])
            worksheet.append(["15.05.2026", "r12399", "r12380", "not_run", "not_available", "", "tool could not complete"])
            workbook.save(workbook_path)

            payload = read_delivery_checklist(profile_id="G65", workspace=trunk)

        checks = {item["key"]: item for item in payload["checks"]}
        self.assertEqual(checks["ramses_size"]["status"], "not_run")
        self.assertEqual(checks["ramses_size"]["raw_value"], "not_run")
        self.assertEqual(checks["logic_size"]["status"], "not_available")
        self.assertEqual(checks["logic_size"]["raw_value"], "not_available")
        self.assertIn("Ramses Size not run", payload["summary"])
        self.assertIn("Logic Size not available", payload["summary"])
        self.assertFalse(payload["is_approval"])

    def test_resolve_delivery_checklist_workbook_supports_mini_filename(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            workbook_path = root / "repositories" / "trunk" / "Cars" / "MINI" / "MINI Export Size.xlsx"
            _write_delivery_workbook(workbook_path)

            resolved = resolve_delivery_checklist_workbook(workspace=root, brand="Mini")

        self.assertEqual(resolved, workbook_path.resolve())

    def test_delivery_checklist_markdown_starts_with_read_only_banner(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "Delivery Data - BMW.xlsx"
            _write_delivery_workbook(workbook_path)
            payload = read_delivery_checklist(
                profile_id="G65",
                workbook_path=workbook_path,
                workspace=Path(temp_dir),
            )

        markdown = render_delivery_checklist_markdown(payload)

        self.assertTrue(markdown.startswith(READ_ONLY_BANNER))
        self.assertIn("SGFX does not run the delivery checklist or modify the workbook.", markdown)
        self.assertIn("Export Size", markdown)
        self.assertIn("Screenshots", markdown)
        self.assertNotIn("approved", markdown.lower())

    def test_daily_digest_surfaces_delivery_checklist_as_evidence_guidance(self) -> None:
        checklist_payload = {
            "profile_id": "G65",
            "status": "available",
            "data_available": True,
            "last_tested": "2026-05-14 08:30",
            "svn_revision": "r12345",
            "changelog_revision": "r12340",
            "checks": [
                {"key": "export_size", "label": "Export Size", "status": "passed"},
                {"key": "screenshots", "label": "Screenshots", "status": "failed"},
            ],
            "summary": "Delivery checklist G65: Export Size passed; Screenshots failed.",
        }

        digest = build_daily_digest(
            {
                "ticket_id": "IDCEVODEV-977874",
                "scope": ["G65"],
                "daily_snapshot_summary": {"smoke_completed": 0, "smoke_total": 0},
                "screenshot_battery_counts": {"total": 0},
                "daily_delta_summary": {},
                "daily_delta": {},
                "review_owner_decisions": {"sections": []},
                "manual_review_profiles": [],
                "delivery_checklist": [checklist_payload],
                "artifact_references": {},
                "top_review_priority_items": [],
                "open_items": [],
            }
        )

        evidence_items = digest["sections"]["evidence_prepared"]["items"]
        delivery_items = [item for item in evidence_items if item.get("source") == "delivery_checklist"]

        self.assertEqual(len(delivery_items), 1)
        self.assertEqual(delivery_items[0]["label"], "Delivery checklist G65")
        self.assertEqual(delivery_items[0]["status"], "prepared")
        self.assertIn("Export Size passed", delivery_items[0]["detail"])
        self.assertIn("Screenshots failed", delivery_items[0]["detail"])
        self.assertIn("evidence guidance", delivery_items[0]["note"])
        self.assertNotIn("approved", delivery_items[0]["note"].lower())
        self.assertIn("Delivery checklist G65", digest["markdown"])

    def test_daily_digest_does_not_hide_unavailable_delivery_checklist_data(self) -> None:
        checklist_payload = {
            "profile_id": "G65",
            "status": "unavailable",
            "data_available": False,
            "summary": "delivery-checklist data unavailable: workbook not found.",
            "workbook_path": "C:/missing/Delivery Data - BMW.xlsx",
        }

        digest = build_daily_digest(
            {
                "ticket_id": "IDCEVODEV-977874",
                "scope": ["G65"],
                "daily_snapshot_summary": {"smoke_completed": 0, "smoke_total": 0},
                "screenshot_battery_counts": {"total": 0},
                "daily_delta_summary": {},
                "daily_delta": {},
                "review_owner_decisions": {"sections": []},
                "manual_review_profiles": [],
                "delivery_checklist": [checklist_payload],
                "artifact_references": {},
                "top_review_priority_items": [],
                "open_items": [],
            }
        )

        evidence_items = digest["sections"]["evidence_prepared"]["items"]
        delivery_items = [item for item in evidence_items if item.get("source") == "delivery_checklist"]

        self.assertEqual(len(delivery_items), 1)
        self.assertEqual(delivery_items[0]["status"], "unavailable")
        self.assertIn("data unavailable", delivery_items[0]["detail"])
        self.assertIn("Delivery checklist G65", digest["markdown"])


if __name__ == "__main__":
    unittest.main()
