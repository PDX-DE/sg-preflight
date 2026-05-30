from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook

from sg_preflight.bmw_delivery import candidate_bmw_profile_ids, resolve_svn_profile_id


READ_ONLY_BANNER = (
    "Export-size analysis data is read-only from operator-local size_analysis workbooks. "
    "SGFX does not run the export size workflow or modify the workbook."
)
_NOTE = "Read-only export-size analysis evidence guidance; not approval or delivery signoff."
_OVERVIEW_SHEET = "Overview"


def _workspace_root(workspace: Path | str | None = None) -> Path:
    root = Path(workspace) if workspace is not None else Path(__file__).resolve().parents[1]
    return root.resolve()


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="minutes")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, float):
        return format(value, ".12g")
    return str(value).strip()


def _profile_token(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.casefold())


def _candidate_profile_ids(profile_id: str) -> tuple[str, ...]:
    candidates: list[str] = []
    for item in (resolve_svn_profile_id(profile_id), *candidate_bmw_profile_ids(profile_id)):
        if item and item not in candidates:
            candidates.append(item)
        if item.upper().endswith("_EVO"):
            stripped = item[:-4]
            if stripped and stripped not in candidates:
                candidates.append(stripped)
    normalized = profile_id.strip().upper()
    if normalized and normalized not in candidates:
        candidates.append(normalized)
    if normalized.endswith("_EVO"):
        stripped = normalized[:-4]
        if stripped and stripped not in candidates:
            candidates.append(stripped)
    return tuple(candidates)


def _candidate_profile_tokens(profile_id: str) -> set[str]:
    return {_profile_token(item) for item in _candidate_profile_ids(profile_id) if _profile_token(item)}


def _size_analysis_dirs(root: Path) -> tuple[Path, ...]:
    return (
        root / "Cars" / "size_analysis",
        root / "repositories" / "trunk" / "Cars" / "size_analysis",
        root / "size_analysis",
    )


def _date_token(value: str | None) -> str:
    if not value:
        return ""
    digits = re.sub(r"[^0-9]", "", str(value))
    return digits if len(digits) == 8 else ""


def _date_text_from_token(value: str) -> str:
    token = _date_token(value)
    if not token:
        return ""
    return f"{token[0:4]}-{token[4:6]}-{token[6:8]}"


def _filename_date_token(path: Path) -> str:
    match = re.search(r"_(\d{8})$", path.stem)
    return match.group(1) if match else ""


def _workbook_metadata(workbook_path: Path, *, row_count: int = 0, workbook_profile: str = "") -> dict[str, Any]:
    stat = workbook_path.stat() if workbook_path.exists() else None
    return {
        "file_size": int(stat.st_size) if stat else 0,
        "modified_at": datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds") if stat else "",
        "row_count": row_count,
        "workbook_profile": workbook_profile,
    }


def _missing_payload(
    profile_id: str,
    workbook_path: Path,
    status: str,
    summary: str,
    *,
    workbook_date: str = "",
    workbook_profile: str = "",
) -> dict[str, Any]:
    return {
        "profile_id": profile_id.strip(),
        "matched_profile_id": workbook_profile,
        "status": status,
        "data_available": False,
        "workbook_path": str(workbook_path),
        "workbook_date": workbook_date,
        "worksheet": "",
        "workbook_metadata": _workbook_metadata(workbook_path, workbook_profile=workbook_profile),
        "variant_count": 0,
        "variants": [],
        "summary": summary,
        "note": _NOTE,
        "is_approval": False,
    }


def _candidate_paths(root: Path, profile_id: str, date: str | None = None) -> tuple[Path, ...]:
    token = _date_token(date)
    suffix = token if token else "*"
    paths: list[Path] = []
    for directory in _size_analysis_dirs(root):
        for profile in _candidate_profile_ids(profile_id):
            paths.append(directory / f"{profile}_{suffix}.xlsx")
    return tuple(paths)


def _find_latest_workbook(root: Path, profile_id: str) -> Path | None:
    matches: list[tuple[str, Path]] = []
    seen: set[Path] = set()
    for directory in _size_analysis_dirs(root):
        if not directory.exists():
            continue
        for profile in _candidate_profile_ids(profile_id):
            for path in directory.glob(f"{profile}_*.xlsx"):
                resolved = path.resolve()
                if resolved in seen:
                    continue
                seen.add(resolved)
                date_token = _filename_date_token(path)
                if date_token:
                    matches.append((date_token, resolved))
    if not matches:
        return None
    return sorted(matches, key=lambda item: (item[0], str(item[1])))[-1][1]


def resolve_export_size_analysis_workbook(
    *,
    profile_id: str,
    workspace: Path | str | None = None,
    workbook_path: Path | str | None = None,
    date: str | None = None,
    latest: bool = False,
) -> Path:
    if workbook_path is not None:
        return Path(workbook_path).resolve()
    root = _workspace_root(workspace)
    if latest or not date:
        latest_workbook = _find_latest_workbook(root, profile_id)
        if latest_workbook is not None:
            return latest_workbook
    candidates = _candidate_paths(root, profile_id, date)
    for candidate in candidates:
        if candidate.exists():
            return candidate.resolve()
    for candidate in candidates:
        if candidate.parent.exists():
            return candidate.resolve()
    return candidates[0].resolve()


def _find_overview_header(rows: list[tuple[object, ...]]) -> tuple[int, list[str]]:
    for index, row in enumerate(rows):
        headers = [_cell_text(value) for value in row]
        normalized = [re.sub(r"[^a-z0-9]+", "", item.casefold()) for item in headers]
        if normalized and normalized[0] == "variant" and "total" in normalized:
            return index, headers
    return -1, []


def _overview_profile(rows: list[tuple[object, ...]], workbook: Path) -> str:
    for row in rows:
        first_cell = _cell_text(row[0] if row else "")
        if first_cell and first_cell.casefold() != "variant":
            return first_cell
    stem = re.sub(r"_\d{8}$", "", workbook.stem)
    return stem


def _variant_rows(rows: list[tuple[object, ...]], header_index: int, headers: list[str]) -> list[dict[str, Any]]:
    variants: list[dict[str, Any]] = []
    for row in rows[header_index + 1 :]:
        name = _cell_text(row[0] if row else "")
        if not name:
            continue
        raw_row = [_cell_text(value) for value in row[: len(headers)]]
        totals: dict[str, str] = {}
        for index, header in enumerate(headers[1:], start=1):
            label = header.strip()
            if not label:
                continue
            value = _cell_text(row[index] if index < len(row) else "")
            totals[label] = value
        variants.append({"name": name, "totals": totals, "raw_row": raw_row})
    return variants


def read_export_size_analysis(
    *,
    profile_id: str,
    workspace: Path | str | None = None,
    workbook_path: Path | str | None = None,
    date: str | None = None,
    latest: bool = False,
) -> dict[str, Any]:
    workbook = resolve_export_size_analysis_workbook(
        profile_id=profile_id,
        workspace=workspace,
        workbook_path=workbook_path,
        date=date,
        latest=latest,
    )
    profile = profile_id.strip()
    workbook_date = _date_text_from_token(_filename_date_token(workbook) or _date_token(date))
    if not workbook.exists():
        return _missing_payload(
            profile,
            workbook,
            "unavailable",
            (
                f"export-size analysis data unavailable: workbook not found for {profile or 'profile'}: {workbook}. "
                "BMW export may be complete, but workbook generation is a CI team operation."
            ),
            workbook_date=workbook_date,
        )

    try:
        loaded = load_workbook(workbook, read_only=True, data_only=True)
    except Exception as exc:
        return _missing_payload(
            profile,
            workbook,
            "unreadable",
            f"export-size analysis data unavailable: workbook could not be read: {exc}",
            workbook_date=workbook_date,
        )

    try:
        if _OVERVIEW_SHEET not in loaded.sheetnames:
            return _missing_payload(
                profile,
                workbook,
                "no_overview_sheet",
                f"export-size analysis data unavailable: Overview sheet was not found in {workbook}.",
                workbook_date=workbook_date,
            )
        worksheet = loaded[_OVERVIEW_SHEET]
        rows = [tuple(row) for row in worksheet.iter_rows(values_only=True)]
        workbook_profile = _overview_profile(rows, workbook)
        header_index, headers = _find_overview_header(rows)
        if header_index < 0:
            return _missing_payload(
                profile,
                workbook,
                "no_overview_sheet",
                f"export-size analysis data unavailable: Overview sheet did not contain the expected Variant header.",
                workbook_date=workbook_date,
                workbook_profile=workbook_profile,
            )
        if workbook_profile and _profile_token(workbook_profile) not in _candidate_profile_tokens(profile):
            return _missing_payload(
                profile,
                workbook,
                "profile_not_found",
                f"export-size analysis data unavailable: workbook profile {workbook_profile} did not match {profile}.",
                workbook_date=workbook_date,
                workbook_profile=workbook_profile,
            )
        variants = _variant_rows(rows, header_index, headers)
    finally:
        loaded.close()

    matched_profile = workbook_profile or re.sub(r"_\d{8}$", "", workbook.stem)
    date_part = f" {workbook_date}" if workbook_date else ""
    summary = (
        f"Export-size analysis {matched_profile}{date_part}: "
        f"{len(variants)} variant{'s' if len(variants) != 1 else ''} recorded from Overview sheet."
    )
    return {
        "profile_id": profile,
        "matched_profile_id": matched_profile,
        "status": "available",
        "data_available": True,
        "workbook_path": str(workbook),
        "workbook_date": workbook_date,
        "worksheet": _OVERVIEW_SHEET,
        "workbook_metadata": _workbook_metadata(
            workbook,
            row_count=len(rows),
            workbook_profile=matched_profile,
        ),
        "variant_count": len(variants),
        "variants": variants,
        "summary": summary,
        "note": _NOTE,
        "is_approval": False,
    }


def read_export_size_analyses_for_profiles(
    profile_ids: list[str] | tuple[str, ...],
    *,
    workspace: Path | str | None = None,
    workbook_path: Path | str | None = None,
) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    for profile_id in profile_ids:
        payload = read_export_size_analysis(
            profile_id=profile_id,
            workspace=workspace,
            workbook_path=workbook_path,
            latest=True,
        )
        items.append(payload)
    return items


def export_size_analysis_digest_items(state: dict[str, Any]) -> list[dict[str, Any]]:
    raw_items = state.get("export_size_analysis", [])
    if isinstance(raw_items, dict):
        raw_items = [raw_items]
    if not isinstance(raw_items, list):
        return []
    items: list[dict[str, Any]] = []
    for raw_item in raw_items:
        if not isinstance(raw_item, dict):
            continue
        profile = str(raw_item.get("matched_profile_id") or raw_item.get("profile_id") or "profile").strip()
        workbook_date = str(raw_item.get("workbook_date", "")).strip()
        variants = raw_item.get("variants", [])
        variant_names = [
            str(item.get("name", "")).strip()
            for item in variants
            if isinstance(item, dict) and str(item.get("name", "")).strip()
        ]
        variant_count = int(raw_item.get("variant_count", len(variant_names)) or 0)
        sample = ", ".join(variant_names[:3])
        detail_parts = [f"{variant_count} variant{'s' if variant_count != 1 else ''}"]
        if sample:
            detail_parts.append(f"sample: {sample}")
        detail = "; ".join(detail_parts)
        if not raw_item.get("data_available"):
            detail = str(raw_item.get("summary", detail)).strip()
        date_suffix = f" ({workbook_date})" if workbook_date else ""
        items.append(
            {
                "label": f"Export-size analysis {profile}{date_suffix}",
                "status": "prepared" if raw_item.get("data_available") else str(raw_item.get("status", "not_available")),
                "detail": detail,
                "source": "export_size_analysis",
                "path": str(raw_item.get("workbook_path", "")).strip(),
                "note": _NOTE,
            }
        )
    return items


def render_export_size_analysis_markdown(payload: dict[str, Any]) -> str:
    lines = [
        READ_ONLY_BANNER,
        "",
        f"# Export-size Analysis Evidence - {payload.get('profile_id', 'profile')}",
        "",
        f"- Status: `{payload.get('status', 'unknown')}`",
        f"- Data available: `{str(bool(payload.get('data_available'))).lower()}`",
    ]
    workbook_path = str(payload.get("workbook_path", "")).strip()
    if workbook_path:
        lines.append(f"- Workbook: `{workbook_path}`")
    workbook_date = str(payload.get("workbook_date", "")).strip()
    if workbook_date:
        lines.append(f"- Workbook date: `{workbook_date}`")
    metadata = payload.get("workbook_metadata", {})
    if isinstance(metadata, dict):
        modified_at = str(metadata.get("modified_at", "")).strip()
        row_count = int(metadata.get("row_count", 0) or 0)
        if modified_at:
            lines.append(f"- Workbook modified: `{modified_at}`")
        if row_count:
            lines.append(f"- Rows scanned: `{row_count}`")
    variant_count = int(payload.get("variant_count", 0) or 0)
    lines.append(f"- Variants recorded: `{variant_count}`")
    summary = str(payload.get("summary", "")).strip()
    if summary:
        lines.extend(["", summary])
    variants = payload.get("variants", [])
    if isinstance(variants, list) and variants:
        lines.extend(["", "## Overview Variants"])
        for variant in variants:
            if not isinstance(variant, dict):
                continue
            name = str(variant.get("name", "variant")).strip()
            totals = variant.get("totals", {})
            if isinstance(totals, dict):
                total_parts = [
                    f"{label}: `{value}`"
                    for label, value in totals.items()
                    if str(value).strip()
                ]
            else:
                total_parts = []
            suffix = f" - {'; '.join(total_parts)}" if total_parts else ""
            lines.append(f"- {name}{suffix}")
    lines.extend(["", "Manual delivery review remains required."])
    return "\n".join(lines).rstrip() + "\n"


def render_export_size_analysis_text(payload: dict[str, Any]) -> str:
    lines = [
        READ_ONLY_BANNER,
        str(payload.get("summary", "Export-size analysis status unavailable.")),
        "Manual delivery review remains required.",
    ]
    workbook_path = str(payload.get("workbook_path", "")).strip()
    if workbook_path:
        lines.append(f"Workbook: {workbook_path}")
    variants = payload.get("variants", [])
    if isinstance(variants, list):
        for variant in variants:
            if not isinstance(variant, dict):
                continue
            lines.append(f"- {variant.get('name', 'variant')}")
    return "\n".join(lines)
