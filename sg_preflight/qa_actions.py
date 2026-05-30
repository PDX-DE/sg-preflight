from __future__ import annotations

import json
import os
import re
import shutil
import subprocess
import sys
import uuid
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl

from sg_preflight.checker_evidence import (
    merge_checker_evidence,
    parse_delivery_checklist_log,
    parse_repo_checker_outputs,
    parse_scene_check_output,
    parse_unused_resources_output,
)
from sg_preflight.profiles import RunProfile, list_run_profiles, resolve_source_repo_root
from sg_preflight.services import (
    RunRequest,
    build_progress_payload,
    execute_profile_run,
    operator_ui_runs_root,
    prerequisite_status,
    utc_now,
    workspace_root,
    write_json_file,
)
from sg_preflight.visual_review import materialize_visual_review_prep


ACTION_PROGRESS_PLANS: dict[str, tuple[tuple[str, str], ...]] = {
    "daily_live_matrix": (
        ("queued", "Queued"),
        ("profiles", "Run live profiles"),
        ("finalize", "Finalize shared summary"),
    ),
    "profile_stack": (
        ("queued", "Queued"),
        ("preflight", "Run standard preflight"),
        ("repo_checker", "Run repo checker"),
        ("unused_resources", "Run unused resource scan"),
        ("scene_check", "Run scene check"),
        ("delivery_checklist", "Check delivery checklist readiness"),
        ("bmw_smoke", "Check BMW smoke readiness"),
        ("finalize", "Finalize action record"),
    ),
    "repo_checker": (
        ("queued", "Queued"),
        ("style", "Run style checker"),
        ("execute", "Run checker"),
        ("parse", "Parse checker output"),
        ("finalize", "Finalize action record"),
    ),
    "bmw_screenshot_smoke": (
        ("queued", "Queued"),
        ("export", "Run BMW export"),
        ("screenshots", "Run BMW screenshots"),
        ("finalize", "Finalize action record"),
    ),
    "scene_check": (
        ("queued", "Queued"),
        ("discover", "Discover scenes"),
        ("execute", "Run scene checks"),
        ("finalize", "Finalize action record"),
    ),
    "unused_resources": (
        ("queued", "Queued"),
        ("execute", "Run scan"),
        ("parse", "Parse scan output"),
        ("finalize", "Finalize action record"),
    ),
    "delivery_checklist": (
        ("queued", "Queued"),
        ("inspect", "Inspect checklist bridge"),
        ("summarize", "Summarize readiness"),
        ("finalize", "Finalize action record"),
    ),
}


def operator_ui_actions_root(explicit_root: Path | None = None) -> Path:
    return workspace_root(explicit_root) / "out" / "operator-ui" / "actions"


def _repo_checker_paths(mirror_root: Path) -> tuple[Path, Path]:
    checkers_root = mirror_root / ".pdx" / "checkers"
    return (
        checkers_root / "code_style_checker" / "check_all_styles.py",
        checkers_root / "executeChecks.py",
    )


def _repo_checker_command_preview(style_script: Path, checker_script: Path, target: Path) -> str:
    return (
        f"{sys.executable} {style_script} {target} && "
        f"{sys.executable} {checker_script} {target}"
    )


def _unused_resources_script_path(mirror_root: Path) -> Path:
    return mirror_root / ".pdx" / "checkers" / "printNotUsedResources.py"


def _unused_resources_inputs(project_root: Path) -> tuple[Path, Path]:
    return project_root / "resources", project_root


def _unused_resources_command_preview(script: Path, project_root: Path) -> str:
    resources_root, rca_root = _unused_resources_inputs(project_root)
    return f"{sys.executable} {script} --res {resources_root} --rca {rca_root}"


def _delivery_checklist_paths(mirror_root: Path) -> dict[str, Path]:
    checklist_root = mirror_root / ".pdx" / "checkers" / "deliveryChecklist"
    return {
        "root": checklist_root,
        "tool": checklist_root / "deliveryChecklist.exe",
        "helper": checklist_root / "deliveryChecklist.py",
        "readme": checklist_root / "README.md",
        "camera_crane": checklist_root / "cameraCrane.lua",
    }


def _delivery_checklist_command_preview(profile: RunProfile) -> str:
    return (
        "internal: inspect mirrored deliveryChecklist assets and BMW-side prerequisites "
        f"for {profile.profile_id}"
    )


def _repo_checker_target(record: ActionRecord, mirror_root: Path) -> Path:
    if record.project_root:
        return Path(record.project_root)
    if record.action_id == "repo_checker_all":
        return mirror_root
    if record.action_id == "repo_checker_idcevo":
        return mirror_root / "Cars_IDCevo"
    return mirror_root / "Cars"


@dataclass(frozen=True)
class OperatorAction:
    action_id: str
    label: str
    description: str
    kind: str
    scope: str
    ready: bool
    blocker_message: str = ""
    profile_id: str = ""
    project_root: str = ""
    command_preview: str = ""

    def to_dict(self) -> dict[str, Any]:
        return {
            "action_id": self.action_id,
            "label": self.label,
            "description": self.description,
            "kind": self.kind,
            "scope": self.scope,
            "ready": self.ready,
            "blocker_message": self.blocker_message,
            "profile_id": self.profile_id,
            "project_root": self.project_root,
            "command_preview": self.command_preview,
        }


@dataclass
class ActionRecord:
    run_id: str
    action_id: str
    label: str
    kind: str
    scope: str
    status: str
    created_at_utc: str
    started_at_utc: str | None
    completed_at_utc: str | None
    workspace_root: str
    profile_id: str = ""
    project_root: str = ""
    command_preview: str = ""
    blocker_message: str = ""
    error_message: str = ""
    exit_code: int | None = None
    paths: dict[str, str] = field(default_factory=dict)
    artifacts: list[dict[str, str]] = field(default_factory=list)
    manual_evidence: list[dict[str, str]] = field(default_factory=list)
    summary: dict[str, Any] | None = None
    notes: list[str] = field(default_factory=list)
    progress: dict[str, Any] | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "schema_version": 1,
            "run_id": self.run_id,
            "action_id": self.action_id,
            "label": self.label,
            "kind": self.kind,
            "scope": self.scope,
            "status": self.status,
            "created_at_utc": self.created_at_utc,
            "started_at_utc": self.started_at_utc,
            "completed_at_utc": self.completed_at_utc,
            "workspace_root": self.workspace_root,
            "profile_id": self.profile_id,
            "project_root": self.project_root,
            "command_preview": self.command_preview,
            "blocker_message": self.blocker_message,
            "error_message": self.error_message,
            "exit_code": self.exit_code,
            "paths": dict(self.paths),
            "artifacts": [dict(item) for item in self.artifacts],
            "manual_evidence": [dict(item) for item in self.manual_evidence],
            "summary": self.summary,
            "notes": list(self.notes),
            "progress": dict(self.progress) if isinstance(self.progress, dict) else None,
        }

    @classmethod
    def from_dict(cls, payload: dict[str, Any]) -> "ActionRecord":
        artifacts = payload.get("artifacts", [])
        return cls(
            run_id=str(payload.get("run_id", "")),
            action_id=str(payload.get("action_id", "")),
            label=str(payload.get("label", "")),
            kind=str(payload.get("kind", "")),
            scope=str(payload.get("scope", "")),
            status=str(payload.get("status", "")),
            created_at_utc=str(payload.get("created_at_utc", "")),
            started_at_utc=payload.get("started_at_utc"),
            completed_at_utc=payload.get("completed_at_utc"),
            workspace_root=str(payload.get("workspace_root", "")),
            profile_id=str(payload.get("profile_id", "")),
            project_root=str(payload.get("project_root", "")),
            command_preview=str(payload.get("command_preview", "")),
            blocker_message=str(payload.get("blocker_message", "")),
            error_message=str(payload.get("error_message", "")),
            exit_code=payload.get("exit_code"),
            paths=dict(payload.get("paths", {}))
            if isinstance(payload.get("paths"), dict)
            else {},
            artifacts=[dict(item) for item in artifacts if isinstance(item, dict)],
            manual_evidence=[dict(item) for item in payload.get("manual_evidence", []) if isinstance(item, dict)],
            summary=payload.get("summary") if isinstance(payload.get("summary"), dict) else None,
            notes=[str(item) for item in payload.get("notes", []) if item],
            progress=dict(payload.get("progress", {}))
            if isinstance(payload.get("progress"), dict)
            else None,
        )


def _status_map(root: Path) -> dict[str, dict[str, str]]:
    return {item["key"]: item for item in prerequisite_status(root)}


def _path_from_status(status_map: dict[str, dict[str, str]], key: str) -> Path:
    raw = status_map.get(key, {}).get("path", "")
    return Path(raw) if raw else Path()


def _status_value(status_map: dict[str, dict[str, str]], key: str) -> str:
    return str(status_map.get(key, {}).get("status", "")).strip().lower()


def _status_detail(status_map: dict[str, dict[str, str]], key: str) -> str:
    return str(status_map.get(key, {}).get("detail", "")).strip()


def _scene_runtime_blocker_message(status_map: dict[str, dict[str, str]]) -> str:
    raco_status = _status_value(status_map, "raco_headless")
    raco_detail = _status_detail(status_map, "raco_headless")
    if raco_status == "incompatible":
        detail = f" {raco_detail}" if raco_detail else ""
        return "Scene check is blocked because the configured `RaCoHeadless.exe` cannot open the representative SG scene." + detail
    return "Scene check needs both the mirrored `check_scenes.py` helper and a locally compatible `RaCoHeadless.exe`."


def _bmw_smoke_script_path(status_map: dict[str, dict[str, str]], profile: RunProfile) -> Path:
    bmw_repo = _path_from_status(status_map, "bmw_models_repo")
    if not bmw_repo:
        return Path()
    return bmw_repo / "ci" / "scripts" / (profile.bmw_smoke_runner or "car_manager.py")


def _bmw_smoke_blocker_message(status_map: dict[str, dict[str, str]], profile: RunProfile) -> str:
    bmw_repo = _path_from_status(status_map, "bmw_models_repo")
    script_path = _bmw_smoke_script_path(status_map, profile)
    if not bmw_repo.exists():
        return "BMW screenshot smoke needs a local `digital-3d-car-models` clone from BMW Git."
    if not script_path.exists():
        return "BMW screenshot smoke needs the `ci/scripts/car_manager.py` helper in the BMW models repo."
    if not profile.bmw_smoke_target.strip():
        return f"No BMW screenshot-smoke target mapping is configured for {profile.profile_id} yet."
    return ""


def list_operator_actions(
    workspace: Path | None = None,
    *,
    profiles: list[RunProfile] | None = None,
) -> list[OperatorAction]:
    root = workspace_root(workspace)
    live_profiles = profiles or list_run_profiles(root)
    status_map = _status_map(root)
    source_root = resolve_source_repo_root(root)
    mirror_root = root / "repositories" / "trunk"
    style_script, checker_script = _repo_checker_paths(mirror_root)
    unused_resources_script = _unused_resources_script_path(mirror_root)
    delivery_checklist_paths = _delivery_checklist_paths(mirror_root)
    scene_checker = mirror_root / "check_scenes.py"
    raco_headless = _path_from_status(status_map, "raco_headless")
    checker_ready = style_script.exists() and checker_script.exists()
    scene_ready = scene_checker.exists() and raco_headless.exists() and _status_value(status_map, "raco_headless") == "available"

    actions = [
        OperatorAction(
            action_id="repo_checker_all",
            label="Run full repo checkers",
            description="Run the SG checker stack over the live SG repo root, matching `checkall.bat` scope without calling the batch wrapper directly.",
            kind="repo_checker",
            scope="workspace",
            ready=checker_ready and source_root.exists(),
            blocker_message=(
                ""
                if checker_ready and source_root.exists()
                else "The SG checker stack (`check_all_styles.py` + `executeChecks.py`) or live source repo root is missing."
            ),
            command_preview=_repo_checker_command_preview(
                style_script,
                checker_script,
                source_root,
            ),
        ),
        OperatorAction(
            action_id="daily_live_matrix",
            label="Run daily SG check",
            description="Run the recommended SG QA stack across the ready live SG slices and write one shared summary.",
            kind="daily_live_matrix",
            scope="workspace",
            ready=any(profile.source_project_root().exists() and profile.config_path.exists() for profile in live_profiles),
            blocker_message=(
                ""
                if any(profile.source_project_root().exists() and profile.config_path.exists() for profile in live_profiles)
                else "No ready live SG profiles are configured on this machine."
            ),
            command_preview="internal: run the recommended QA stack across all ready live profiles",
        ),
        OperatorAction(
            action_id="repo_checker_idcevo",
            label="Run IDCevo repo checkers",
            description="Run the SG checker stack over the mirrored `Cars_IDCevo` tree.",
            kind="repo_checker",
            scope="workspace",
            ready=checker_ready and (source_root / "Cars_IDCevo").exists(),
            blocker_message=(
                ""
                if checker_ready and (source_root / "Cars_IDCevo").exists()
                else "The SG checker stack (`check_all_styles.py` + `executeChecks.py`) or live `Cars_IDCevo` tree is missing."
            ),
            command_preview=_repo_checker_command_preview(
                style_script,
                checker_script,
                source_root / "Cars_IDCevo",
            ),
        ),
        OperatorAction(
            action_id="repo_checker_classic",
            label="Run classic repo checkers",
            description="Run the SG checker stack over the mirrored `Cars` tree.",
            kind="repo_checker",
            scope="workspace",
            ready=checker_ready and (source_root / "Cars").exists(),
            blocker_message=(
                ""
                if checker_ready and (source_root / "Cars").exists()
                else "The SG checker stack (`check_all_styles.py` + `executeChecks.py`) or live `Cars` tree is missing."
            ),
            command_preview=_repo_checker_command_preview(
                style_script,
                checker_script,
                source_root / "Cars",
            ),
        ),
    ]

    for profile in live_profiles:
        source_project_root = profile.source_project_root()
        actions.append(
            OperatorAction(
                action_id=f"qa_stack__{profile.profile_id.lower()}",
                label=f"Run recommended QA stack for {profile.profile_id}",
                description=(
                    f"Run the default preflight first, then every additional SG-side QA step that is available on this machine for {profile.profile_id}."
                ),
                kind="profile_stack",
                scope="profile",
                ready=source_project_root.exists() and profile.config_path.exists(),
                blocker_message=(
                    ""
                    if source_project_root.exists() and profile.config_path.exists()
                    else f"The project root or config for {profile.profile_id} is missing, so the recommended stack cannot start."
                ),
                profile_id=profile.profile_id,
                project_root=str(source_project_root),
                command_preview=(
                    "internal: standard preflight + repo checker + unused resource scan + scene check + delivery checklist readiness + BMW smoke readiness summary"
                ),
            )
        )
        actions.append(
            OperatorAction(
                action_id=f"repo_checker_profile__{profile.profile_id.lower()}",
                label=f"Run repo check for {profile.profile_id}",
                description=f"Run the SG checker stack only for the {profile.profile_id} project tree.",
                kind="repo_checker",
                scope="profile",
                ready=checker_ready and source_project_root.exists(),
                blocker_message=(
                    ""
                    if checker_ready and source_project_root.exists()
                    else (
                        f"The SG checker stack (`check_all_styles.py` + `executeChecks.py`) "
                        f"or project root for {profile.profile_id} is missing."
                    )
                ),
                profile_id=profile.profile_id,
                project_root=str(source_project_root),
                command_preview=_repo_checker_command_preview(
                    style_script,
                    checker_script,
                    source_project_root,
                ),
            )
        )
        actions.append(
            OperatorAction(
                action_id=f"unused_resources__{profile.profile_id.lower()}",
                label=f"Run unused resource scan for {profile.profile_id}",
                description=(
                    f"Run the SG unused-resource checker for the {profile.profile_id} project so leftover resource files can be reviewed before handoff."
                ),
                kind="unused_resources",
                scope="profile",
                ready=(
                    unused_resources_script.exists()
                    and source_project_root.exists()
                    and _unused_resources_inputs(source_project_root)[0].exists()
                    and any(source_project_root.rglob("*.rca"))
                ),
                blocker_message=(
                    ""
                    if (
                        unused_resources_script.exists()
                        and source_project_root.exists()
                        and _unused_resources_inputs(source_project_root)[0].exists()
                        and any(source_project_root.rglob("*.rca"))
                    )
                    else (
                        "Unused resource scan needs `printNotUsedResources.py`, a local `resources` tree, and at least one `.rca` scene under the project root."
                    )
                ),
                profile_id=profile.profile_id,
                project_root=str(source_project_root),
                command_preview=_unused_resources_command_preview(
                    unused_resources_script,
                    source_project_root,
                ),
            )
        )
        delivery_checklist_ready = source_project_root.exists() and all(
            path.exists()
            for key, path in delivery_checklist_paths.items()
            if key != "root"
        )
        actions.append(
            OperatorAction(
                action_id=f"delivery_checklist__{profile.profile_id.lower()}",
                label=f"Check delivery checklist readiness for {profile.profile_id}",
                description=(
                    f"Inspect the SG delivery-checklist bridge assets plus BMW-side prerequisites for {profile.profile_id} without pretending the external BMW flow runs here."
                ),
                kind="delivery_checklist",
                scope="profile",
                ready=delivery_checklist_ready,
                blocker_message=(
                    ""
                    if delivery_checklist_ready
                    else (
                        "Delivery checklist readiness needs the mirrored `.pdx/checkers/deliveryChecklist` assets "
                        "(`deliveryChecklist.exe`, `deliveryChecklist.py`, `README.md`, and `cameraCrane.lua`)."
                    )
                ),
                profile_id=profile.profile_id,
                project_root=str(source_project_root),
                command_preview=_delivery_checklist_command_preview(profile),
            )
        )
        actions.append(
            OperatorAction(
                action_id=f"scene_check__{profile.profile_id.lower()}",
                label=f"Run scene check for {profile.profile_id}",
                description=f"Run SG scene checking over every `.rca` under the {profile.profile_id} project tree.",
                kind="scene_check",
                scope="profile",
                ready=scene_ready and source_project_root.exists(),
                blocker_message=(
                    ""
                    if scene_ready and source_project_root.exists()
                    else (
                        "The mirrored `check_scenes.py` helper is missing."
                        if not scene_checker.exists()
                        else _scene_runtime_blocker_message(status_map)
                    )
                ),
                profile_id=profile.profile_id,
                project_root=str(source_project_root),
                command_preview=f"{sys.executable} {scene_checker} --raco {raco_headless} --dir {source_project_root}",
            )
        )
        bmw_smoke_blocker = _bmw_smoke_blocker_message(status_map, profile)
        bmw_script = _bmw_smoke_script_path(status_map, profile)
        target = profile.bmw_smoke_target.strip()
        actions.append(
            OperatorAction(
                action_id=f"bmw_screenshot_smoke__{profile.profile_id.lower()}",
                label=f"Run BMW screenshot smoke for {profile.profile_id}",
                description=(
                    f"Run BMW-side export and screenshot smoke for {profile.profile_id} when the BMW models repo and car mapping are available."
                ),
                kind="bmw_screenshot_smoke",
                scope="profile",
                ready=not bmw_smoke_blocker,
                blocker_message=bmw_smoke_blocker,
                profile_id=profile.profile_id,
                project_root=str(source_project_root),
                command_preview=(
                    f"{sys.executable} {bmw_script} export {target} && "
                    f"{sys.executable} {bmw_script} screenshots --diff {target}"
                    if target
                    else "BMW screenshot smoke target mapping is not configured yet."
                ),
            )
        )

    return actions


def get_operator_action(
    action_id: str,
    workspace: Path | None = None,
    *,
    profiles: list[RunProfile] | None = None,
) -> OperatorAction:
    normalized = action_id.strip().lower()
    for action in list_operator_actions(workspace, profiles=profiles):
        if action.action_id.lower() == normalized:
            return action
    supported = ", ".join(action.action_id for action in list_operator_actions(workspace, profiles=profiles))
    raise KeyError(f"Unsupported action {action_id!r}. Supported actions: {supported}")


def _default_action_run_id(action_id: str) -> str:
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    return f"{stamp}-{action_id.lower()}-{uuid.uuid4().hex[:6]}"


def build_action_record(action: OperatorAction, workspace: Path | None = None) -> ActionRecord:
    root = workspace_root(workspace)
    run_id = _default_action_run_id(action.action_id)
    output_root = operator_ui_actions_root(root) / run_id
    return ActionRecord(
        run_id=run_id,
        action_id=action.action_id,
        label=action.label,
        kind=action.kind,
        scope=action.scope,
        status="queued",
        created_at_utc=utc_now(),
        started_at_utc=None,
        completed_at_utc=None,
        workspace_root=str(root),
        profile_id=action.profile_id,
        project_root=action.project_root,
        command_preview=action.command_preview,
        blocker_message=action.blocker_message,
        paths={
            "output_root": str(output_root),
            "run_record": str(output_root / "action.json"),
            "action_record": str(output_root / "action.json"),
            "log": str(output_root / "action.log"),
            "summary_json": str(output_root / "summary.json"),
            "summary_md": str(output_root / "summary.md"),
            "xlsx_report": str(output_root / "scene-check.xlsx"),
            "manual_evidence_root": str(output_root / "manual-evidence"),
            "manual_evidence_index": str(output_root / "manual-evidence" / "attachments.json"),
        },
    )


def save_action_record(record: ActionRecord) -> None:
    write_json_file(Path(record.paths["run_record"]), record.to_dict())


def _manual_evidence_root(record: ActionRecord) -> Path:
    configured = str(record.paths.get("manual_evidence_root", "")).strip()
    if configured:
        return Path(configured)
    return Path(record.paths["output_root"]) / "manual-evidence"


def _manual_evidence_index_path(record: ActionRecord) -> Path:
    configured = str(record.paths.get("manual_evidence_index", "")).strip()
    if configured:
        return Path(configured)
    return _manual_evidence_root(record) / "attachments.json"


def _manual_evidence_slug(value: str) -> str:
    slug = re.sub(r"[^a-z0-9._-]+", "-", value.strip().lower())
    slug = slug.strip("-._")
    return slug or "manual-evidence"


def _manual_evidence_default_label(kind: str, source_path: Path | None = None) -> str:
    if kind == "screenshot":
        return "Local manual screenshot evidence"
    if kind == "blender_note":
        return "Blender review note"
    if kind == "raco_note":
        return "RaCo review note"
    if kind == "visual_review_checklist":
        return "Visual review checklist"
    if kind == "verification_note":
        return "Manual verification note"
    if kind == "external_file" and source_path is not None:
        return source_path.name
    return "Manual evidence"


def _manual_evidence_default_note(kind: str, label: str) -> str:
    templates = {
        "blender_note": (
            f"{label}\n\n"
            "- Area checked:\n"
            "- What matched in Blender:\n"
            "- What still needs SG / RaCo follow-up:\n"
        ),
        "raco_note": (
            f"{label}\n\n"
            "- Scene checked:\n"
            "- What matched in RaCo:\n"
            "- What still needs SG follow-up:\n"
        ),
        "visual_review_checklist": (
            f"{label}\n\n"
            "- Project changelog reviewed: [ ]\n"
            "- Screenshot baseline set reviewed: [ ]\n"
            "- Blender scene opened: [ ]\n"
            "- RaCo scene opened: [ ]\n"
            "- Blender vs RaCo compared: [ ]\n"
            "- Key camera / perspective checked: [ ]\n"
            "- Screenshot captured: [ ]\n"
            "- Constants / README notes checked: [ ]\n"
            "- Finding documented: [ ]\n"
            "- Notes:\n"
        ),
        "verification_note": (
            f"{label}\n\n"
            "- What was verified:\n"
            "- Result:\n"
            "- Remaining blocker or follow-up:\n"
        ),
    }
    return templates.get(kind, f"{label}\n")


def _manual_evidence_target_path(record: ActionRecord, kind: str, label: str, source_path: Path | None = None) -> Path:
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    suffix = source_path.suffix if source_path is not None and source_path.suffix else ".md"
    basename = f"{stamp}-{_manual_evidence_slug(kind)}-{_manual_evidence_slug(label)}-{uuid.uuid4().hex[:6]}{suffix}"
    return _manual_evidence_root(record) / basename


def _write_manual_evidence_index(record: ActionRecord) -> None:
    write_json_file(
        _manual_evidence_index_path(record),
        {
            "schema_version": 1,
            "run_id": record.run_id,
            "items": [dict(item) for item in record.manual_evidence],
        },
    )


def attach_manual_evidence(
    run_id_or_path: str | Path,
    workspace: Path | None = None,
    *,
    kind: str,
    label: str = "",
    source_path: str = "",
    note: str = "",
) -> dict[str, str]:
    root = workspace_root(workspace)
    record = load_action_record(run_id_or_path, root)
    source = Path(source_path).expanduser() if source_path.strip() else None
    if source is not None and not source.exists():
        raise FileNotFoundError(f"Manual evidence source was not found: {source}")

    manual_root = _manual_evidence_root(record)
    manual_root.mkdir(parents=True, exist_ok=True)

    resolved_label = label.strip() or _manual_evidence_default_label(kind, source)
    resolved_note = note.strip()
    target_path = _manual_evidence_target_path(record, kind, resolved_label, source)

    if kind in {"blender_note", "raco_note", "visual_review_checklist", "verification_note"}:
        text = resolved_note or _manual_evidence_default_note(kind, resolved_label)
        target_path.write_text(text.strip() + "\n", encoding="utf-8")
        resolved_note = text.strip()
    elif source is not None:
        source_resolved = source.resolve()
        target_resolved = target_path.resolve()
        if source_resolved != target_resolved:
            shutil.copy2(source_resolved, target_resolved)
        else:
            target_path = source_resolved
    else:
        raise ValueError("Manual evidence attachment requires a source file or a note-based kind.")

    entry = {
        "id": uuid.uuid4().hex,
        "kind": kind.strip(),
        "label": resolved_label,
        "path": str(target_path),
        "note": resolved_note,
        "source_path": str(source.resolve()) if source is not None else "",
        "created_at_utc": utc_now(),
    }
    record.manual_evidence.append(entry)
    record.artifacts.append(_artifact(resolved_label, target_path))
    _write_manual_evidence_index(record)
    save_action_record(record)
    return entry


def _progress_event(
    step_key: str,
    label: str,
    detail: str = "",
    meta: dict[str, Any] | None = None,
) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "timestamp_utc": utc_now(),
        "step_key": step_key,
        "label": label,
        "detail": detail,
    }
    if meta:
        payload["meta"] = dict(meta)
    return payload


def _merged_progress_events(
    existing: dict[str, Any] | None,
    *,
    step_key: str,
    label: str,
    detail: str = "",
    meta: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    raw_events = existing.get("events", []) if isinstance(existing, dict) else []
    events = [dict(item) for item in raw_events if isinstance(item, dict)]
    if (
        not events
        or events[-1].get("step_key") != step_key
        or events[-1].get("label") != label
        or events[-1].get("detail") != detail
        or (
            isinstance(meta, dict)
            and dict(events[-1].get("meta", {})) != dict(meta)
        )
    ):
        events.append(_progress_event(step_key, label, detail, meta))
    return events[-60:]


def _set_action_progress(
    record: ActionRecord,
    *,
    step_key: str,
    percent: int,
    label: str,
    detail: str = "",
    meta: dict[str, Any] | None = None,
) -> None:
    plan = ACTION_PROGRESS_PLANS.get(record.kind, (("queued", "Queued"), ("finalize", "Finalize action record")))
    events = _merged_progress_events(
        record.progress,
        step_key=step_key,
        label=label,
        detail=detail,
        meta=meta,
    )
    record.progress = build_progress_payload(
        plan,
        step_key=step_key,
        percent=percent,
        label=label,
        detail=detail,
        events=events,
    )
    save_action_record(record)


def load_action_record(path_or_run_id: str | Path, workspace: Path | None = None) -> ActionRecord:
    candidate = Path(path_or_run_id)
    if candidate.exists():
        record_path = candidate if candidate.is_file() else candidate / "action.json"
    else:
        record_path = operator_ui_actions_root(workspace) / str(path_or_run_id) / "action.json"
    payload = json.loads(record_path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise ValueError(f"Action record must contain a JSON object: {record_path}")
    return ActionRecord.from_dict(payload)


def list_recent_action_records(workspace: Path | None = None, limit: int = 12) -> list[ActionRecord]:
    records: list[ActionRecord] = []
    actions_root = operator_ui_actions_root(workspace)
    if not actions_root.exists():
        return records

    for path in actions_root.iterdir():
        record_path = path / "action.json"
        if not record_path.exists():
            continue
        try:
            records.append(load_action_record(record_path, workspace))
        except (OSError, ValueError, json.JSONDecodeError):
            continue

    records.sort(key=lambda item: item.created_at_utc, reverse=True)
    return records[:limit]


def _write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def _artifact(label: str, path: Path) -> dict[str, str]:
    return {"label": label, "path": str(path)}


def _visual_review_prep_entries(record: ActionRecord, root: Path) -> tuple[list[dict[str, str]], list[str]]:
    project_root = Path(record.project_root)
    if not project_root.exists():
        return [], []

    output_root = Path(record.paths.get("output_root", "")).resolve()
    output_root.mkdir(parents=True, exist_ok=True)

    bundle = materialize_visual_review_prep(
        record.profile_id or project_root.name,
        project_root,
        output_root,
        repo_root=resolve_source_repo_root(root),
    )
    record.paths["visual_review_prep_json"] = str(bundle.json_path)
    record.paths["visual_review_prep_md"] = str(bundle.markdown_path)
    record.paths["visual_review_gallery_html"] = str(bundle.html_path)

    prep = bundle.prep
    artifacts = [
        _artifact("Visual review prep", bundle.markdown_path),
        _artifact("Visual review gallery", bundle.html_path),
        _artifact("Visual review prep JSON", bundle.json_path),
    ]
    if prep.changelog_path:
        artifacts.append(_artifact("Project changelog", Path(prep.changelog_path)))
    for readme_path in prep.project_readme_paths[:4]:
        artifacts.append(_artifact("Project README", Path(readme_path)))
    if prep.screenshot_root:
        artifacts.append(_artifact("Screenshot baselines", Path(prep.screenshot_root)))
    if prep.screenshot_test_config_path:
        artifacts.append(_artifact("Screenshot test config", Path(prep.screenshot_test_config_path)))
    if prep.constants_readme_path:
        artifacts.append(_artifact("Constants README", Path(prep.constants_readme_path)))
    if prep.raco_scene_path:
        artifacts.append(_artifact("Representative RaCo scene", Path(prep.raco_scene_path)))
    if prep.blender_workfile_path:
        artifacts.append(_artifact("Representative Blender workfile", Path(prep.blender_workfile_path)))
    for shared_doc_path in prep.shared_doc_paths[:6]:
        artifacts.append(_artifact("Shared BMW doc", Path(shared_doc_path)))

    notes: list[str] = []
    if prep.changelog_heading:
        notes.append(f"Visual review focus: {prep.changelog_heading}")
    if prep.project_svn_log_lines:
        notes.append("Recent project SVN: " + " | ".join(prep.project_svn_log_lines[:2]))
    if prep.shared_svn_log_lines:
        notes.append("Recent shared SVN: " + " | ".join(prep.shared_svn_log_lines[:2]))
    if prep.priority_screenshots:
        notes.append("Priority screenshot baselines: " + ", ".join(prep.priority_screenshots[:6]))
    if prep.shared_doc_paths:
        notes.append("Shared BMW docs to review: " + ", ".join(Path(path).name for path in prep.shared_doc_paths[:4]))
    if prep.raco_scene_path:
        notes.append(f"Representative RaCo scene ready to open: {prep.raco_scene_path}")
    if prep.blender_workfile_path:
        notes.append(f"Representative Blender workfile ready to open: {prep.blender_workfile_path}")
    notes.append("Open the visual review gallery and review-prep note before closing the manual visual-review step.")
    return artifacts, notes


def _summary_checker_evidence(summary: dict[str, Any] | None) -> dict[str, Any] | None:
    if not isinstance(summary, dict):
        return None
    evidence = summary.get("checker_evidence")
    return evidence if isinstance(evidence, dict) else None


def aggregate_child_checker_evidence(
    *summaries: dict[str, Any] | None,
    raw_log_path: str = "",
    source_kind: str = "profile_stack",
) -> dict[str, Any] | None:
    evidence_items = [
        evidence
        for evidence in (_summary_checker_evidence(summary) for summary in summaries)
        if evidence is not None
    ]
    if not evidence_items:
        return None
    merged = merge_checker_evidence(*evidence_items)
    merged["raw_log_path"] = raw_log_path
    merged["source_kind"] = source_kind
    return merged


def _nested_action_record(parent: ActionRecord, action: OperatorAction, slug: str) -> ActionRecord:
    run_id = f"{parent.run_id}-{slug}"
    output_root = operator_ui_actions_root(Path(parent.workspace_root)) / run_id
    return ActionRecord(
        run_id=run_id,
        action_id=action.action_id,
        label=action.label,
        kind=action.kind,
        scope=action.scope,
        status="running",
        created_at_utc=parent.created_at_utc,
        started_at_utc=utc_now(),
        completed_at_utc=None,
        workspace_root=parent.workspace_root,
        profile_id=action.profile_id,
        project_root=action.project_root,
        command_preview=action.command_preview,
        blocker_message=action.blocker_message,
        paths={
            "output_root": str(output_root),
            "run_record": str(output_root / "action.json"),
            "action_record": str(output_root / "action.json"),
            "log": str(output_root / "action.log"),
            "summary_json": str(output_root / "summary.json"),
            "summary_md": str(output_root / "summary.md"),
            "xlsx_report": str(output_root / "scene-check.xlsx"),
            "manual_evidence_root": str(output_root / "manual-evidence"),
            "manual_evidence_index": str(output_root / "manual-evidence" / "attachments.json"),
        },
    )


def _complete_nested_action_record(
    record: ActionRecord,
    summary: dict[str, Any],
    artifacts: list[dict[str, str]],
    notes: list[str],
) -> None:
    record.summary = summary
    record.artifacts = artifacts
    record.notes = notes
    record.status = "completed"
    record.completed_at_utc = utc_now()
    record.exit_code = 0
    record.progress = build_progress_payload(
        ACTION_PROGRESS_PLANS.get(record.kind, (("queued", "Queued"), ("finalize", "Finalize action record"))),
        step_key="finalize",
        percent=100,
        label="Action completed",
        detail="The generated files and summary are ready to open.",
        events=_merged_progress_events(
            record.progress,
            step_key="finalize",
            label="Action completed",
            detail="The generated files and summary are ready to open.",
        ),
    )
    _save_action_summary(record)
    save_action_record(record)


def _log(record: ActionRecord, text: str) -> None:
    log_path = Path(record.paths["log"])
    log_path.parent.mkdir(parents=True, exist_ok=True)
    with log_path.open("a", encoding="utf-8") as handle:
        handle.write(text.rstrip() + "\n")


def _summary_md_lines(summary: dict[str, Any]) -> list[str]:
    lines = [
        f"# {summary.get('title', 'SG Preflight QA Action')}",
        "",
    ]
    for item in summary.get("lines", []):
        lines.append(f"- {item}")
    return lines


def _save_action_summary(record: ActionRecord) -> None:
    if record.summary is None:
        return
    write_json_file(Path(record.paths["summary_json"]), record.summary)
    _write_text(Path(record.paths["summary_md"]), "\n".join(_summary_md_lines(record.summary)).strip() + "\n")


def _parse_repo_checker_output(output: str) -> dict[str, Any]:
    phase_matches = re.findall(r"starting\s+(\w+)\s+on\s+(\d+)\s+files", output, flags=re.IGNORECASE)
    error_matches = [int(value) for value in re.findall(r"(\d+)\s+errors found", output, flags=re.IGNORECASE)]
    phases = [f"{name}: {count} file(s)" for name, count in phase_matches]
    error_total = sum(error_matches)
    lines = []
    if phases:
        lines.extend(phases)
    lines.append(f"Reported error batches: {error_total}")
    return {
        "title": "Repo checker result",
        "lines": lines,
        "phase_count": len(phase_matches),
        "reported_error_batches": error_total,
    }


def _parse_style_checker_output(output: str) -> dict[str, Any]:
    checked_match = re.search(
        r"checked\s+(\d+)\s+files\s+\(src:\s*(\d+);\s*fmt:\s*(\d+);\s*license:\s*(\d+)\)",
        output,
        flags=re.IGNORECASE,
    )
    issue_match = re.search(r"detected\s+(\d+)\s+style guide issues", output, flags=re.IGNORECASE)
    clean = "no style guide violations detected" in output.lower()
    warning_lines = re.findall(r"warning\s+PRJ9999", output, flags=re.IGNORECASE)

    checked_files = int(checked_match.group(1)) if checked_match else 0
    issue_count = int(issue_match.group(1)) if issue_match else (0 if clean else len(warning_lines))
    src_files = int(checked_match.group(2)) if checked_match else 0
    formatting_files = int(checked_match.group(3)) if checked_match else 0
    license_files = int(checked_match.group(4)) if checked_match else 0

    lines = []
    if checked_match:
        lines.append(
            "Style checker scope: "
            f"{checked_files} file(s) checked "
            f"(src: {src_files}, formatting: {formatting_files}, license: {license_files})."
        )
    if clean:
        lines.append("Style checker: no style-guide issues reported.")
    else:
        lines.append(f"Style checker: {issue_count} style-guide issue(s) reported.")

    return {
        "title": "Style checker result",
        "lines": lines,
        "checked_files": checked_files,
        "src_files": src_files,
        "formatting_files": formatting_files,
        "license_files": license_files,
        "issue_count": issue_count,
        "clean": clean,
    }


def _parse_unused_resources_output(output: str, project_root: Path) -> dict[str, Any]:
    unused_files = []
    for raw_line in output.splitlines():
        candidate = raw_line.strip()
        if not candidate:
            continue
        if candidate.startswith("Traceback") or candidate.lower().startswith("usage:"):
            continue
        try:
            path = Path(candidate)
        except OSError:
            continue
        if not path.suffix:
            continue
        try:
            display = str(path.resolve().relative_to(project_root.resolve())).replace("\\", "/")
        except (OSError, ValueError):
            display = str(path)
        unused_files.append(display)

    lines = [f"Unused resources reported: {len(unused_files)}"]
    if unused_files:
        preview = ", ".join(unused_files[:5])
        lines.append(f"First unused resources: {preview}")
    else:
        lines.append("No unused resources were reported in the scanned resource tree.")

    return {
        "title": "Unused resource scan result",
        "lines": lines,
        "unused_count": len(unused_files),
        "unused_files": unused_files,
    }


def _scene_error_blocks(output: str) -> list[str]:
    errors: list[str] = []
    current = ""
    collecting = False
    for line in output.splitlines():
        if line.startswith("[E]"):
            if current:
                errors.append(current)
            current = line
            collecting = True
            continue
        if collecting and (line.startswith("[") or not line.strip()):
            if current:
                errors.append(current)
            current = ""
            collecting = False
            continue
        if collecting:
            current += "\n" + line
    if current:
        errors.append(current)
    return errors


def _attach_scene_workbook_refs(
    checker_evidence: dict[str, Any],
    workbook_refs: dict[str, list[dict[str, Any]]],
) -> dict[str, Any]:
    affected = checker_evidence.get("affected_files", [])
    if not isinstance(affected, list):
        return checker_evidence

    refs_by_path = {
        path: [dict(item) for item in refs]
        for path, refs in workbook_refs.items()
    }
    for item in affected:
        if not isinstance(item, dict):
            continue
        path = str(item.get("path", "")).strip()
        if not path or path not in refs_by_path or not refs_by_path[path]:
            continue
        ref = refs_by_path[path].pop(0)
        item["workbook_sheet"] = ref.get("workbook_sheet")
        item["workbook_row"] = ref.get("workbook_row")
    return checker_evidence


def _execute_daily_live_matrix(record: ActionRecord, root: Path) -> tuple[dict[str, Any], list[dict[str, str]], list[str]]:
    profiles = [
        profile
        for profile in list_run_profiles(root)
        if profile.source_project_root().exists() and profile.config_path.exists()
    ]
    lines = []
    artifacts: list[dict[str, str]] = []
    notes: list[str] = []
    total_profiles = max(len(profiles), 1)
    child_summaries: list[dict[str, Any]] = []

    _set_action_progress(
        record,
        step_key="profiles",
        percent=10,
        label="Running live profile matrix",
        detail=f"Preparing {len(profiles)} live profile run(s).",
    )

    for index, profile in enumerate(profiles, start=1):
        matrix_percent = 10 + int((index - 1) / total_profiles * 75)
        stack_action = get_operator_action(
            f"qa_stack__{profile.profile_id.lower()}",
            root,
            profiles=profiles,
        )
        child_record = _nested_action_record(
            record,
            stack_action,
            f"{profile.profile_id.lower()}-qa-stack",
        )
        _set_action_progress(
            record,
            step_key="profiles",
            percent=matrix_percent,
            label=f"Running {profile.profile_id}",
            detail=f"Live matrix {index}/{len(profiles)}: running the recommended SG QA stack for {profile.profile_id}.",
            meta={
                "profile_id": profile.profile_id,
                "index": index,
                "total": len(profiles),
                "child_run_id": child_record.run_id,
                "child_status_url": f"/ui/api/actions/{child_record.run_id}",
                "child_result_url": f"/ui/actions/{child_record.run_id}",
                "command": child_record.command_preview,
            },
        )
        child_summary, child_artifacts, child_notes = _execute_profile_stack(child_record, root)
        _complete_nested_action_record(
            child_record,
            child_summary,
            child_artifacts,
            child_notes,
        )
        child_summaries.append(child_summary)
        preflight_errors = int(child_summary.get("preflight_errors", 0) or 0)
        preflight_warnings = int(child_summary.get("preflight_warnings", 0) or 0)
        repo_style_issues = int(child_summary.get("repo_style_issue_count", 0) or 0)
        repo_error_batches = int(child_summary.get("repo_execute_error_batches", 0) or 0)
        unused_candidates = int(child_summary.get("unused_candidate_count", 0) or 0)
        scene_errors = int(child_summary.get("scene_error_count", 0) or 0)
        lines.append(
            f"{profile.profile_id}: preflight {preflight_errors} error(s), {preflight_warnings} warning(s); "
            f"repo checker {repo_style_issues} style issue(s) / {repo_error_batches} executeChecks batch(es); "
            f"unused resources {unused_candidates}; scene errors {scene_errors}"
        )
        notes.extend(f"{profile.profile_id}: {note}" for note in child_notes[:3])
        artifacts.extend(child_artifacts)
        artifacts.extend(
            [
                _artifact(f"{profile.profile_id} QA stack summary", Path(child_record.paths["summary_md"])),
                _artifact(f"{profile.profile_id} QA stack JSON", Path(child_record.paths["summary_json"])),
                _artifact(f"{profile.profile_id} QA stack log", Path(child_record.paths["log"])),
            ]
        )
        _log(record, f"{profile.profile_id}: completed with {child_summary}")

    _set_action_progress(
        record,
        step_key="finalize",
        percent=92,
        label="Finalizing daily matrix",
        detail="Writing the shared SG live-matrix summary and artifacts.",
    )
    checker_evidence = aggregate_child_checker_evidence(
        *child_summaries,
        raw_log_path=record.paths["log"],
        source_kind="daily_live_matrix",
    )
    if checker_evidence is not None and checker_evidence.get("top_paths"):
        first_path = checker_evidence["top_paths"][0]
        first_line = f" line {first_path['line']}" if first_path.get("line") not in (None, "") else ""
        lines.append(
            f"Open first: {first_path['path']}{first_line} ({first_path.get('checker', 'checker')}) - {first_path.get('message', '')}"
        )
    summary = {
        "title": "Daily SG check",
        "lines": lines,
        "profile_count": len(profiles),
    }
    if checker_evidence is not None:
        summary["checker_evidence"] = checker_evidence
    return summary, artifacts, notes


def _execute_profile_stack(record: ActionRecord, root: Path) -> tuple[dict[str, Any], list[dict[str, str]], list[str]]:
    profile = next(
        candidate
        for candidate in list_run_profiles(root)
        if candidate.profile_id.lower() == record.profile_id.lower()
    )
    lines: list[str] = []
    artifacts: list[dict[str, str]] = []
    notes: list[str] = []

    _set_action_progress(
        record,
        step_key="preflight",
        percent=12,
        label="Running standard preflight",
        detail=f"Starting the recommended deterministic check for {profile.profile_id}.",
        meta={
            "profile_id": profile.profile_id,
            "child_run_id": f"{record.run_id}-preflight",
            "child_status_url": f"/ui/api/runs/{record.run_id}-preflight",
            "child_result_url": f"/ui/runs/{record.run_id}-preflight",
        },
    )
    preflight_run_id = f"{record.run_id}-preflight"
    preflight_output = operator_ui_runs_root(root) / preflight_run_id
    preflight = execute_profile_run(
        profile,
        RunRequest(
            profile_id=profile.profile_id,
            fail_on="never",
            output_root=preflight_output,
            run_id=preflight_run_id,
        ),
        root,
    )
    preflight_summary = preflight.summary or {}
    lines.append(
        "Standard preflight: "
        f"{preflight_summary.get('errors', 0)} errors, "
        f"{preflight_summary.get('warnings', 0)} warnings, "
        f"{preflight_summary.get('info', 0)} info"
    )
    artifacts.extend(
        [
            _artifact("Standard preflight HTML report", Path(preflight.paths["html_report"])),
            _artifact("Standard preflight Markdown report", Path(preflight.paths["markdown_report"])),
            _artifact("Standard preflight JSON report", Path(preflight.paths["json_report"])),
            _artifact("Standard preflight run record", Path(preflight.paths["run_record"])),
        ]
    )
    notes.extend(f"Preflight: {note}" for note in preflight.notes[:3])
    repo_summary: dict[str, Any] | None = None
    unused_summary: dict[str, Any] | None = None
    scene_summary: dict[str, Any] | None = None
    delivery_summary: dict[str, Any] | None = None

    repo_action = get_operator_action(
        f"repo_checker_profile__{profile.profile_id.lower()}",
        root,
        profiles=[profile],
    )
    _set_action_progress(
        record,
        step_key="repo_checker",
        percent=46,
        label="Checking SG repo hygiene",
        detail=f"Running mirrored repo checker coverage for {profile.profile_id}.",
        meta={},
    )
    if repo_action.ready:
        repo_record = _nested_action_record(record, repo_action, "repo-checker")
        _set_action_progress(
            record,
            step_key="repo_checker",
            percent=46,
            label="Checking SG repo hygiene",
            detail=f"Running mirrored repo checker coverage for {profile.profile_id}.",
            meta={
                "child_run_id": repo_record.run_id,
                "child_status_url": f"/ui/api/actions/{repo_record.run_id}",
                "child_result_url": f"/ui/actions/{repo_record.run_id}",
                "command": repo_record.command_preview,
            },
        )
        repo_summary, repo_artifacts, repo_notes = _execute_repo_checker(repo_record, root)
        _complete_nested_action_record(repo_record, repo_summary, repo_artifacts, repo_notes)
        lines.append(
            "Repo checker: "
            f"{repo_summary.get('style_issue_count', 0)} style issue(s), "
            f"{repo_summary.get('reported_error_batches', 0)} executeChecks error batch(es), "
            f"{repo_summary.get('phase_count', 0)} executeChecks phase(s)"
        )
        artifacts.extend(repo_artifacts)
    else:
        lines.append(f"Repo checker: blocked - {repo_action.blocker_message}")

    unused_resources_action = get_operator_action(
        f"unused_resources__{profile.profile_id.lower()}",
        root,
        profiles=[profile],
    )
    _set_action_progress(
        record,
        step_key="unused_resources",
        percent=58,
        label="Checking unused resources",
        detail=f"Running resource-to-scene usage scan for {profile.profile_id}.",
        meta={},
    )
    if unused_resources_action.ready:
        unused_record = _nested_action_record(record, unused_resources_action, "unused-resources")
        _set_action_progress(
            record,
            step_key="unused_resources",
            percent=58,
            label="Checking unused resources",
            detail=f"Running resource-to-scene usage scan for {profile.profile_id}.",
            meta={
                "child_run_id": unused_record.run_id,
                "child_status_url": f"/ui/api/actions/{unused_record.run_id}",
                "child_result_url": f"/ui/actions/{unused_record.run_id}",
                "command": unused_record.command_preview,
            },
        )
        unused_summary, unused_artifacts, unused_notes = _execute_unused_resources(unused_record, root)
        _complete_nested_action_record(unused_record, unused_summary, unused_artifacts, unused_notes)
        lines.append(
            "Unused resources: "
            f"{unused_summary.get('unused_count', 0)} candidate file(s) reported"
        )
        artifacts.extend(unused_artifacts)
    else:
        lines.append(f"Unused resources: blocked - {unused_resources_action.blocker_message}")

    scene_action = get_operator_action(
        f"scene_check__{profile.profile_id.lower()}",
        root,
        profiles=[profile],
    )
    _set_action_progress(
        record,
        step_key="scene_check",
        percent=72,
        label="Checking RaCo scenes",
        detail=f"Running scene-check coverage for {profile.profile_id} where local RaCo is available.",
        meta={},
    )
    if scene_action.ready:
        scene_record = _nested_action_record(record, scene_action, "scene-check")
        _set_action_progress(
            record,
            step_key="scene_check",
            percent=72,
            label="Checking RaCo scenes",
            detail=f"Running scene-check coverage for {profile.profile_id} where local RaCo is available.",
            meta={
                "child_run_id": scene_record.run_id,
                "child_status_url": f"/ui/api/actions/{scene_record.run_id}",
                "child_result_url": f"/ui/actions/{scene_record.run_id}",
                "command": scene_record.command_preview,
            },
        )
        scene_summary, scene_artifacts, scene_notes = _execute_scene_check(scene_record, root)
        _complete_nested_action_record(scene_record, scene_summary, scene_artifacts, scene_notes)
        lines.append(
            "Scene check: "
            f"{scene_summary.get('checked_scenes', 0)} scenes checked, "
            f"{scene_summary.get('scenes_with_errors', 0)} with errors"
        )
        artifacts.extend(scene_artifacts)
    else:
        lines.append(f"Scene check: blocked - {scene_action.blocker_message}")

    delivery_action = get_operator_action(
        f"delivery_checklist__{profile.profile_id.lower()}",
        root,
        profiles=[profile],
    )
    _set_action_progress(
        record,
        step_key="delivery_checklist",
        percent=84,
        label="Checking delivery checklist bridge",
        detail=f"Inspecting mirrored delivery-checklist assets and BMW-side prerequisites for {profile.profile_id}.",
        meta={},
    )
    if delivery_action.ready:
        delivery_record = _nested_action_record(record, delivery_action, "delivery-checklist")
        _set_action_progress(
            record,
            step_key="delivery_checklist",
            percent=84,
            label="Checking delivery checklist bridge",
            detail=f"Inspecting mirrored delivery-checklist assets and BMW-side prerequisites for {profile.profile_id}.",
            meta={
                "child_run_id": delivery_record.run_id,
                "child_status_url": f"/ui/api/actions/{delivery_record.run_id}",
                "child_result_url": f"/ui/actions/{delivery_record.run_id}",
                "command": delivery_record.command_preview,
            },
        )
        delivery_summary, delivery_artifacts, delivery_notes = _execute_delivery_checklist(delivery_record, root)
        _complete_nested_action_record(delivery_record, delivery_summary, delivery_artifacts, delivery_notes)
        lines.append(
            "Delivery checklist: "
            + "; ".join(str(line) for line in delivery_summary.get("lines", [])[:2])
        )
        artifacts.extend(delivery_artifacts)
    else:
        lines.append(f"Delivery checklist: blocked - {delivery_action.blocker_message}")

    bmw_action = get_operator_action(
        f"bmw_screenshot_smoke__{profile.profile_id.lower()}",
        root,
        profiles=[profile],
    )
    _set_action_progress(
        record,
        step_key="bmw_smoke",
        percent=92,
        label="Checking BMW smoke stage",
        detail=f"Evaluating BMW-side screenshot smoke coverage for {profile.profile_id}.",
        meta={},
    )
    if bmw_action.ready:
        bmw_record = _nested_action_record(record, bmw_action, "bmw-screenshot-smoke")
        _set_action_progress(
            record,
            step_key="bmw_smoke",
            percent=92,
            label="Checking BMW smoke stage",
            detail=f"Evaluating BMW-side screenshot smoke coverage for {profile.profile_id}.",
            meta={
                "child_run_id": bmw_record.run_id,
                "child_status_url": f"/ui/api/actions/{bmw_record.run_id}",
                "child_result_url": f"/ui/actions/{bmw_record.run_id}",
                "command": bmw_record.command_preview,
            },
        )
        bmw_summary, bmw_artifacts, _ = _execute_bmw_screenshot_smoke(bmw_record, root)
        lines.extend(f"BMW screenshot smoke: {line}" for line in bmw_summary.get("lines", []))
        artifacts.extend(bmw_artifacts)
    else:
        lines.append(f"BMW screenshot smoke: blocked - {bmw_action.blocker_message}")

    checker_evidence = aggregate_child_checker_evidence(
        repo_summary,
        unused_summary,
        scene_summary,
        delivery_summary,
        raw_log_path=record.paths["log"],
        source_kind="profile_stack",
    )
    if checker_evidence is not None and checker_evidence.get("top_paths"):
        first_path = checker_evidence["top_paths"][0]
        first_line = f" line {first_path['line']}" if first_path.get("line") not in (None, "") else ""
        lines.append(
            f"Open first: {first_path['path']}{first_line} ({first_path.get('checker', 'checker')}) - {first_path.get('message', '')}"
        )
    summary = {
        "title": f"Recommended QA Stack - {profile.profile_id}",
        "lines": lines,
        "profile_id": profile.profile_id,
        "preflight_errors": preflight_summary.get("errors", 0),
        "preflight_warnings": preflight_summary.get("warnings", 0),
        "preflight_info": preflight_summary.get("info", 0),
        "repo_style_issue_count": repo_summary.get("style_issue_count", 0) if isinstance(repo_summary, dict) else 0,
        "repo_execute_error_batches": (
            repo_summary.get("reported_error_batches", 0) if isinstance(repo_summary, dict) else 0
        ),
        "unused_candidate_count": unused_summary.get("unused_count", 0) if isinstance(unused_summary, dict) else 0,
        "scene_checked_scenes": scene_summary.get("checked_scenes", 0) if isinstance(scene_summary, dict) else 0,
        "scene_error_count": scene_summary.get("scenes_with_errors", 0) if isinstance(scene_summary, dict) else 0,
        "delivery_local_assets_found": (
            delivery_summary.get("local_assets_found", 0) if isinstance(delivery_summary, dict) else 0
        ),
        "delivery_bmw_repo_ready": (
            bool(delivery_summary.get("bmw_repo_ready")) if isinstance(delivery_summary, dict) else False
        ),
    }
    if checker_evidence is not None:
        summary["checker_evidence"] = checker_evidence
    return summary, artifacts, notes


def _execute_repo_checker(record: ActionRecord, root: Path) -> tuple[dict[str, Any], list[dict[str, str]], list[str]]:
    mirror_root = root / "repositories" / "trunk"
    source_root = resolve_source_repo_root(root)
    style_script, checker_script = _repo_checker_paths(mirror_root)
    target = _repo_checker_target(record, source_root)
    env = os.environ.copy()
    env["SG-Repo"] = str(source_root)
    _set_action_progress(
        record,
        step_key="style",
        percent=18,
        label="Running SG style checker",
        detail=f"Launching check_all_styles.py on {target}.",
        meta={
            "target": str(target),
            "command": f"{sys.executable} {style_script} {target}",
        },
    )
    style_result = subprocess.run(
        [sys.executable, str(style_script), str(target)],
        cwd=root,
        capture_output=True,
        text=True,
        check=False,
        env=env,
    )
    _set_action_progress(
        record,
        step_key="execute",
        percent=46,
        label="Running SG repo checker",
        detail=f"Launching executeChecks.py on {target}.",
        meta={
            "target": str(target),
            "command": f"{sys.executable} {checker_script} {target}",
        },
    )
    result = subprocess.run(
        [sys.executable, str(checker_script), str(target)],
        cwd=root,
        capture_output=True,
        text=True,
        check=False,
        env=env,
    )
    style_output = ((style_result.stdout or "") + ("\n" + style_result.stderr if style_result.stderr else "")).strip()
    checker_output = ((result.stdout or "") + ("\n" + result.stderr if result.stderr else "")).strip()
    combined_output = "\n\n".join(
        [
            "=== check_all_styles.py ===",
            style_output,
            "",
            "=== executeChecks.py ===",
            checker_output,
        ]
    ).strip()
    _write_text(Path(record.paths["log"]), combined_output + ("\n" if combined_output else ""))
    _set_action_progress(
        record,
        step_key="parse",
        percent=78,
        label="Parsing SG repo checker output",
        detail="Summarizing style, license, Lua, shader, and executeChecks results.",
        meta={"target": str(target)},
    )
    style_summary = _parse_style_checker_output(style_output)
    checker_summary = _parse_repo_checker_output(checker_output)
    checker_evidence = parse_repo_checker_outputs(
        style_output,
        checker_output,
        raw_log_path=record.paths["log"],
    )
    summary = {
        "title": "Repo checker result",
        "lines": [
            f"Style checker: {style_summary.get('issue_count', 0)} style-guide issue(s) across {style_summary.get('checked_files', 0)} checked file(s).",
            f"executeChecks: {checker_summary.get('reported_error_batches', 0)} error batch(es) across {checker_summary.get('phase_count', 0)} phase(s).",
            f"Scope: {target}",
        ],
        "style_issue_count": style_summary.get("issue_count", 0),
        "style_checked_files": style_summary.get("checked_files", 0),
        "style_src_files": style_summary.get("src_files", 0),
        "style_formatting_files": style_summary.get("formatting_files", 0),
        "style_license_files": style_summary.get("license_files", 0),
        "style_return_code": style_result.returncode,
        "phase_count": checker_summary.get("phase_count", 0),
        "reported_error_batches": checker_summary.get("reported_error_batches", 0),
        "execute_return_code": result.returncode,
        "phases": list(checker_summary.get("lines", [])),
        "checker_evidence": checker_evidence,
    }
    phase_lines = [
        line
        for line in checker_summary.get("lines", [])
        if isinstance(line, str) and line.lower().startswith(("luacheck:", "shadercheck:", "tabbingcheck:", "newlinecheck:", "binarycheck:"))
    ]
    if phase_lines:
        summary["lines"].append("Phases: " + "; ".join(phase_lines))
    if checker_evidence.get("top_paths"):
        first_path = checker_evidence["top_paths"][0]
        first_line = f" line {first_path['line']}" if first_path.get("line") not in (None, "") else ""
        summary["lines"].append(
            f"Open first: {first_path['path']}{first_line} ({first_path.get('checker', 'checker')}) - {first_path.get('message', '')}"
        )
    summary["lines"].append(
        f"Exit codes: style={style_result.returncode}, executeChecks={result.returncode}."
    )
    artifacts = [
        _artifact("Checker log", Path(record.paths["log"])),
        _artifact("Style checker script", style_script),
        _artifact("executeChecks script", checker_script),
    ]
    return summary, artifacts, []


def _execute_unused_resources(record: ActionRecord, root: Path) -> tuple[dict[str, Any], list[dict[str, str]], list[str]]:
    mirror_root = root / "repositories" / "trunk"
    script_path = _unused_resources_script_path(mirror_root)
    project_root = Path(record.project_root)
    resources_root, rca_root = _unused_resources_inputs(project_root)

    _set_action_progress(
        record,
        step_key="execute",
        percent=28,
        label="Running unused resource scan",
        detail=f"Scanning {resources_root} against `.rca` files under {rca_root}.",
        meta={
            "resources_root": str(resources_root),
            "rca_root": str(rca_root),
            "command": f"{sys.executable} {script_path} --res {resources_root} --rca {rca_root}",
        },
    )
    result = subprocess.run(
        [sys.executable, str(script_path), "--res", str(resources_root), "--rca", str(rca_root)],
        cwd=root,
        capture_output=True,
        text=True,
        check=False,
    )
    output = ((result.stdout or "") + ("\n" + result.stderr if result.stderr else "")).strip()
    _write_text(Path(record.paths["log"]), output + ("\n" if output else ""))

    _set_action_progress(
        record,
        step_key="parse",
        percent=76,
        label="Parsing unused resource output",
        detail="Summarizing candidate resource files that are not referenced by any scanned scene.",
        meta={"project_root": str(project_root)},
    )
    summary = _parse_unused_resources_output(output, project_root)
    checker_evidence = parse_unused_resources_output(
        output,
        raw_log_path=record.paths["log"],
    )
    summary["checker_evidence"] = checker_evidence
    summary["return_code"] = result.returncode
    if checker_evidence.get("top_paths"):
        first_path = checker_evidence["top_paths"][0]
        summary.setdefault("lines", []).append(
            f"Open first: {first_path['path']} ({first_path.get('checker', 'checker')}) - {first_path.get('message', '')}"
        )
    if result.returncode != 0:
        summary.setdefault("lines", []).append(f"Process returned exit code {result.returncode}.")
    artifacts = [
        _artifact("Unused resource scan log", Path(record.paths["log"])),
        _artifact("Unused resource script", script_path),
    ]
    return summary, artifacts, []


def _delivery_checklist_viewer_candidates(bmw_repo: Path) -> list[Path]:
    if not bmw_repo.exists():
        return []

    matches: list[Path] = []
    seen: set[Path] = set()
    for pattern in ("ramses*viewer*.exe", "Ramses*Viewer*.exe"):
        try:
            for candidate in bmw_repo.rglob(pattern):
                resolved = candidate.resolve()
                if resolved in seen:
                    continue
                seen.add(resolved)
                matches.append(candidate)
                if len(matches) >= 4:
                    return matches
        except OSError:
            break
    return matches


def _execute_delivery_checklist(
    record: ActionRecord,
    root: Path,
) -> tuple[dict[str, Any], list[dict[str, str]], list[str]]:
    mirror_root = root / "repositories" / "trunk"
    checklist_paths = _delivery_checklist_paths(mirror_root)
    status_map = _status_map(root)
    bmw_repo = _path_from_status(status_map, "bmw_models_repo")
    car_manager = _path_from_status(status_map, "bmw_car_manager_script")
    test_main = _path_from_status(status_map, "bmw_test_main_script")
    viewer_candidates = _delivery_checklist_viewer_candidates(bmw_repo)

    _set_action_progress(
        record,
        step_key="inspect",
        percent=34,
        label="Inspecting delivery checklist bridge",
        detail=f"Reading mirrored `.pdx/checkers/deliveryChecklist` assets for {record.profile_id}.",
        meta={
            "delivery_checklist_root": str(checklist_paths["root"]),
            "bmw_repo": str(bmw_repo),
        },
    )

    local_assets = {
        key: path.exists()
        for key, path in checklist_paths.items()
        if key != "root"
    }
    local_found = sum(1 for value in local_assets.values() if value)
    helper_paths = [path for path in (car_manager, test_main) if path.exists()]

    _set_action_progress(
        record,
        step_key="summarize",
        percent=78,
        label="Summarizing delivery checklist readiness",
        detail="Combining SG delivery-checklist bridge assets with BMW-side dependency discovery.",
        meta={
            "local_assets_found": local_found,
            "local_assets_total": len(local_assets),
            "bmw_repo_ready": bmw_repo.exists(),
            "bmw_helper_count": len(helper_paths),
            "viewer_count": len(viewer_candidates),
        },
    )

    helper_labels = []
    if car_manager.exists():
        helper_labels.append("ci/scripts/car_manager.py")
    if test_main.exists():
        helper_labels.append("ci/scripts/test/main.py")

    lines = [
        f"Local SG bridge: {local_found}/{len(local_assets)} mirrored deliveryChecklist asset(s) found.",
        (
            "BMW-side prerequisites: repo available; helper scripts found: " + ", ".join(helper_labels) + "."
            if helper_labels
            else "BMW-side prerequisites: blocked on local `digital-3d-car-models` access."
            if not bmw_repo.exists()
            else "BMW-side prerequisites: repo available, but `ci/scripts/car_manager.py` and `ci/scripts/test/main.py` are not present locally."
        ),
    ]
    if viewer_candidates:
        viewer_lines: list[str] = []
        for candidate in viewer_candidates[:2]:
            try:
                viewer_lines.append(str(candidate.relative_to(bmw_repo)))
            except ValueError:
                viewer_lines.append(str(candidate))
        lines.append("Viewer candidates: " + ", ".join(viewer_lines))
    elif bmw_repo.exists():
        lines.append("Viewer candidates: no local `ramses*viewer*.exe` hit was found in the BMW repo scan.")
    lines.append(
        "Current role: this is a readiness bridge into the BMW-owned delivery checklist, not a replacement for that external flow."
    )

    log_lines = [
        f"delivery_checklist_root={checklist_paths['root']}",
        f"delivery_checklist_tool={checklist_paths['tool']} :: {'available' if checklist_paths['tool'].exists() else 'missing'}",
        f"delivery_checklist_helper={checklist_paths['helper']} :: {'available' if checklist_paths['helper'].exists() else 'missing'}",
        f"delivery_checklist_readme={checklist_paths['readme']} :: {'available' if checklist_paths['readme'].exists() else 'missing'}",
        f"delivery_checklist_camera_crane={checklist_paths['camera_crane']} :: {'available' if checklist_paths['camera_crane'].exists() else 'missing'}",
        f"bmw_models_repo={bmw_repo} :: {'available' if bmw_repo.exists() else 'missing'}",
        f"bmw_car_manager_script={car_manager} :: {'available' if car_manager.exists() else 'missing'}",
        f"bmw_test_main_script={test_main} :: {'available' if test_main.exists() else 'missing'}",
    ]
    if viewer_candidates:
        for candidate in viewer_candidates:
            log_lines.append(f"viewer_candidate={candidate}")
    else:
        log_lines.append("viewer_candidate=<none>")
    _write_text(Path(record.paths["log"]), "\n".join(log_lines) + "\n")
    checker_evidence = parse_delivery_checklist_log(
        "\n".join(log_lines),
        raw_log_path=record.paths["log"],
    )

    summary = {
        "title": "Delivery checklist readiness",
        "lines": lines,
        "local_assets_found": local_found,
        "local_assets_total": len(local_assets),
        "bmw_repo_ready": bmw_repo.exists(),
        "bmw_helpers_found": helper_labels,
        "viewer_count": len(viewer_candidates),
        "checker_evidence": checker_evidence,
    }
    if checker_evidence.get("top_paths"):
        first_path = checker_evidence["top_paths"][0]
        summary["lines"].append(
            f"Open first: {first_path['path']} ({first_path.get('checker', 'checker')}) - {first_path.get('message', '')}"
        )
    artifacts = [
        _artifact("Delivery checklist README", checklist_paths["readme"]),
        _artifact("Delivery checklist helper", checklist_paths["helper"]),
        _artifact("Delivery checklist executable", checklist_paths["tool"]),
        _artifact("Delivery checklist camera crane", checklist_paths["camera_crane"]),
        _artifact("Delivery checklist log", Path(record.paths["log"])),
    ]
    if car_manager.exists():
        artifacts.append(_artifact("BMW car manager helper", car_manager))
    if test_main.exists():
        artifacts.append(_artifact("BMW test main helper", test_main))
    return summary, artifacts, []


def _execute_bmw_screenshot_smoke(record: ActionRecord, root: Path) -> tuple[dict[str, Any], list[dict[str, str]], list[str]]:
    profile = next(
        candidate
        for candidate in list_run_profiles(root)
        if candidate.profile_id.lower() == record.profile_id.lower()
    )
    status_map = _status_map(root)
    bmw_repo = _path_from_status(status_map, "bmw_models_repo")
    script_path = _bmw_smoke_script_path(status_map, profile)
    target = profile.bmw_smoke_target.strip()
    scripts_root = bmw_repo / "ci" / "scripts"

    _set_action_progress(
        record,
        step_key="export",
        percent=28,
        label="Running BMW export",
        detail=f"Launching BMW export for target {target}.",
        meta={
            "target": target,
            "command": f"{sys.executable} {script_path} export {target}",
        },
    )
    export_process = subprocess.run(
        [sys.executable, str(script_path), "export", target],
        cwd=scripts_root,
        capture_output=True,
        text=True,
        check=False,
    )
    _set_action_progress(
        record,
        step_key="screenshots",
        percent=68,
        label="Running BMW screenshots",
        detail=f"Capturing screenshot diff output for target {target}.",
        meta={
            "target": target,
            "command": f"{sys.executable} {script_path} screenshots --diff {target}",
        },
    )
    screenshots_process = subprocess.run(
        [sys.executable, str(script_path), "screenshots", "--diff", target],
        cwd=scripts_root,
        capture_output=True,
        text=True,
        check=False,
    )
    combined_log = "\n\n".join(
        [
            "=== export ===",
            (export_process.stdout or "").strip(),
            (export_process.stderr or "").strip(),
            "=== screenshots ===",
            (screenshots_process.stdout or "").strip(),
            (screenshots_process.stderr or "").strip(),
        ]
    ).strip()
    _write_text(Path(record.paths["log"]), combined_log + ("\n" if combined_log else ""))

    summary = {
        "title": "BMW screenshot smoke result",
        "lines": [
            f"Target: {target}",
            f"Export exit code: {export_process.returncode}",
            f"Screenshot exit code: {screenshots_process.returncode}",
        ],
        "target": target,
        "export_exit_code": export_process.returncode,
        "screenshots_exit_code": screenshots_process.returncode,
    }
    artifacts = [_artifact("BMW screenshot smoke log", Path(record.paths["log"]))]
    return summary, artifacts, []


def _execute_scene_check(record: ActionRecord, root: Path) -> tuple[dict[str, Any], list[dict[str, str]], list[str]]:
    status_map = _status_map(root)
    raco_exe = _path_from_status(status_map, "raco_headless")
    scene_checker = root / "repositories" / "trunk" / "check_scenes.py"
    if _status_value(status_map, "raco_headless") != "available":
        raise RuntimeError(_scene_runtime_blocker_message(status_map))
    project_root = Path(record.project_root)
    scenes = sorted(project_root.rglob("*.rca"))
    total_scenes = max(len(scenes), 1)
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    scene_errors = 0
    log_lines: list[str] = []
    workbook_refs: dict[str, list[dict[str, Any]]] = {}

    _set_action_progress(
        record,
        step_key="discover",
        percent=12,
        label="Discovering scenes",
        detail=f"Found {len(scenes)} `.rca` scene(s) under {project_root}.",
        meta={
            "project_root": str(project_root),
            "scene_count": len(scenes),
        },
    )

    for index, scene in enumerate(scenes, start=1):
        scene_percent = 18 + int((index - 1) / total_scenes * 68)
        _set_action_progress(
            record,
            step_key="execute",
            percent=scene_percent,
            label=f"Checking scene {index}/{len(scenes)}",
            detail=str(scene),
            meta={
                "scene": str(scene),
                "index": index,
                "total": len(scenes),
                "command": f"{raco_exe} -p {scene} -l 3",
            },
        )
        log_lines.append(f"Checking scene: {scene}")
        result = subprocess.run(
            [str(raco_exe), "-p", str(scene), "-l", "3"],
            cwd=root,
            capture_output=True,
            text=True,
            check=False,
        )
        output = ((result.stdout or "") + ("\n" + result.stderr if result.stderr else "")).strip()
        if output:
            log_lines.append(output)
        errors = _scene_error_blocks(output)
        if not errors:
            continue

        scene_errors += 1
        sheet = workbook.create_sheet()
        sheet.append([str(scene)])
        title_cell = sheet.cell(column=1, row=1)
        title_cell.font = openpyxl.styles.Font(size=16, bold=True)
        workbook_refs[str(scene)] = []
        for row_index, item in enumerate(errors, start=2):
            sheet.append([item])
            workbook_refs[str(scene)].append(
                {
                    "workbook_sheet": sheet.title,
                    "workbook_row": row_index,
                }
            )
        sheet.column_dimensions["A"].width = 155
        for cell in sheet["A"]:
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

    scene_log = "\n".join(log_lines).strip() + ("\n" if log_lines else "")
    _write_text(Path(record.paths["log"]), scene_log)
    checker_evidence = parse_scene_check_output(
        scene_log,
        raw_log_path=record.paths["log"],
        workbook_path=record.paths["xlsx_report"],
    )
    artifacts = [
        _artifact("Scene checker log", Path(record.paths["log"])),
        _artifact("Scene checker script", scene_checker),
    ]
    if workbook.sheetnames:
        workbook_path = Path(record.paths["xlsx_report"])
        workbook_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(workbook_path)
        artifacts.append(_artifact("Scene checker workbook", workbook_path))
        checker_evidence = _attach_scene_workbook_refs(checker_evidence, workbook_refs)
    workbook.close()

    summary = {
        "title": "Scene check result",
        "lines": [
            f"Checked scenes: {len(scenes)}",
            f"Scenes with errors: {scene_errors}",
        ],
        "checked_scenes": len(scenes),
        "scenes_with_errors": scene_errors,
        "checker_evidence": checker_evidence,
    }
    if checker_evidence.get("top_paths"):
        first_path = checker_evidence["top_paths"][0]
        summary["lines"].append(
            f"Open first: {first_path['path']} - {first_path.get('message', '')}"
        )
    return summary, artifacts, []


def execute_operator_action(
    action: OperatorAction,
    workspace: Path | None = None,
    *,
    record: ActionRecord | None = None,
) -> ActionRecord:
    root = workspace_root(workspace)
    record = record or build_action_record(action, root)
    _set_action_progress(
        record,
        step_key="queued",
        percent=0,
        label="Queued locally",
        detail=f"Preparing {action.label}.",
    )

    if not action.ready:
        record.status = "blocked"
        record.completed_at_utc = utc_now()
        record.exit_code = 0
        record.summary = {
            "title": action.label,
            "lines": [action.blocker_message or "This action is blocked on the current machine."],
        }
        record.progress = {
            **build_progress_payload(
                ACTION_PROGRESS_PLANS.get(action.kind, (("queued", "Queued"),)),
                step_key="queued",
                percent=100,
                label="Blocked on this machine",
                detail=action.blocker_message or "This action is blocked on the current machine.",
                events=_merged_progress_events(
                    record.progress,
                    step_key="queued",
                    label="Blocked on this machine",
                    detail=action.blocker_message or "This action is blocked on the current machine.",
                ),
            ),
            "state": "blocked",
        }
        _save_action_summary(record)
        save_action_record(record)
        return record

    record.status = "running"
    record.started_at_utc = utc_now()
    _set_action_progress(
        record,
        step_key="queued",
        percent=4,
        label="Starting automation",
        detail=f"Launching {action.label}.",
    )

    try:
        if action.kind == "daily_live_matrix":
            summary, artifacts, notes = _execute_daily_live_matrix(record, root)
        elif action.kind == "profile_stack":
            summary, artifacts, notes = _execute_profile_stack(record, root)
        elif action.kind == "repo_checker":
            summary, artifacts, notes = _execute_repo_checker(record, root)
        elif action.kind == "unused_resources":
            summary, artifacts, notes = _execute_unused_resources(record, root)
        elif action.kind == "delivery_checklist":
            summary, artifacts, notes = _execute_delivery_checklist(record, root)
        elif action.kind == "bmw_screenshot_smoke":
            summary, artifacts, notes = _execute_bmw_screenshot_smoke(record, root)
        elif action.kind == "scene_check":
            summary, artifacts, notes = _execute_scene_check(record, root)
        else:
            raise ValueError(f"Unsupported action kind: {action.kind}")

        review_artifacts, review_notes = _visual_review_prep_entries(record, root)
        if review_artifacts:
            artifacts.extend(review_artifacts)
        if review_notes:
            notes.extend(review_notes)

        record.summary = summary
        record.artifacts = artifacts
        record.notes = notes
        record.status = "completed"
        record.completed_at_utc = utc_now()
        record.exit_code = 0
        record.progress = build_progress_payload(
            ACTION_PROGRESS_PLANS.get(action.kind, (("queued", "Queued"),)),
            step_key="finalize",
            percent=100,
            label="Action completed",
            detail="The generated files and summary are ready to open.",
            events=_merged_progress_events(
                record.progress,
                step_key="finalize",
                label="Action completed",
                detail="The generated files and summary are ready to open.",
            ),
        )
        _save_action_summary(record)
        save_action_record(record)
        return record
    except Exception as exc:
        record.status = "failed"
        record.error_message = str(exc)
        record.exit_code = 1
        record.completed_at_utc = utc_now()
        existing_progress = dict(record.progress or {})
        failure_step = str(existing_progress.get("step_key", "finalize")).strip() or "finalize"
        events = _merged_progress_events(
            existing_progress,
            step_key=failure_step,
            label="Action failed",
            detail=str(exc),
        )
        record.progress = build_progress_payload(
            ACTION_PROGRESS_PLANS.get(action.kind, (("queued", "Queued"),)),
            step_key=failure_step,
            percent=int(existing_progress.get("percent", 100) or 100),
            label="Action failed",
            detail=str(exc),
            events=events,
        )
        save_action_record(record)
        raise
