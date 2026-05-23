from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook

from sg_preflight.bmw_delivery import candidate_bmw_profile_ids, resolve_svn_profile_id


WORKBOOK_NAME = "Delivery Data - BMW.xlsx"
READ_ONLY_BANNER = (
    "Delivery checklist data is read-only from the operator-local Excel workbook. "
    "SGFX does not run the delivery checklist or modify the workbook."
)
_CHECK_COLUMNS = {
    "export_size": "Export Size",
    "screenshots": "Screenshots",
    "interface": "Interface",
    "perspectives": "Perspectives",
}
_PROFILE_KEYS = {"profile_id", "car", "car_model", "model", "vehicle", "baureihe"}
_HEADER_ALIASES = {
    "profile": "profile_id",
    "profileid": "profile_id",
    "car": "car",
    "carmodel": "car_model",
    "model": "model",
    "vehicle": "vehicle",
    "baureihe": "baureihe",
    "lasttested": "last_tested",
    "lasttest": "last_tested",
    "tested": "last_tested",
    "date": "last_tested",
    "timestamp": "last_tested",
    "svnrevision": "svn_revision",
    "svnrev": "svn_revision",
    "svn": "svn_revision",
    "changelogrevision": "changelog_revision",
    "changelogrev": "changelog_revision",
    "changelog": "changelog_revision",
    "exportsize": "export_size",
    "export": "export_size",
    "screenshots": "screenshots",
    "screenshot": "screenshots",
    "interface": "interface",
    "perspectives": "perspectives",
    "perspective": "perspectives",
    "ramsessize": "ramses_size",
    "ramses": "ramses_size",
    "logicsize": "logic_size",
    "logic": "logic_size",
    "comment": "comment",
    "comments": "comment",
}


def _workspace_root(workspace: Path | str | None = None) -> Path:
    root = Path(workspace) if workspace is not None else Path(__file__).resolve().parents[1]
    return root.resolve()


def _brand_label(brand: str | None) -> str:
    value = str(brand or "BMW").strip()
    return value.upper() if value.casefold() in {"bmw", "mini"} else value.title()


def _workbook_name_for_brand(brand: str | None) -> str:
    return f"Delivery Data - {_brand_label(brand)}.xlsx"


def _export_size_workbook_name_for_brand(brand: str | None) -> str:
    return f"{_brand_label(brand)} Export Size.xlsx"


def _candidate_profile_ids(profile_id: str) -> tuple[str, ...]:
    candidates: list[str] = []
    for item in (resolve_svn_profile_id(profile_id), *candidate_bmw_profile_ids(profile_id)):
        normalized = item.strip().upper()
        if normalized and normalized not in candidates:
            candidates.append(normalized)
        if normalized.endswith("_EVO"):
            stripped = normalized[:-4]
            if stripped and stripped not in candidates:
                candidates.append(stripped)
    normalized = profile_id.strip().upper()
    if normalized and normalized not in candidates:
        candidates.append(normalized)
    if normalized.endswith("_EVO"):
        stripped = normalized[:-4]
        if stripped and stripped not in candidates:
            candidates.append(stripped)
    return tuple(candidates)


def _size_analysis_dirs(root: Path) -> tuple[Path, ...]:
    return (
        root / "Cars" / "size_analysis",
        root / "repositories" / "trunk" / "Cars" / "size_analysis",
        root / "size_analysis",
    )


def _filename_date_token(path: Path) -> str:
    match = re.search(r"_(\d{8})$", path.stem)
    return match.group(1) if match else ""


def _size_analysis_sort_key(path: Path) -> tuple[int, int, str]:
    date_token = _filename_date_token(path)
    if date_token:
        return (3, int(date_token), path.name.casefold())
    suffix_match = re.search(r"_([^_]+)$", path.stem)
    suffix = suffix_match.group(1).casefold() if suffix_match else ""
    version_match = re.fullmatch(r"v(\d+)", suffix)
    if version_match:
        return (2, int(version_match.group(1)), path.name.casefold())
    if suffix == "vx":
        return (2, 1_000_000, path.name.casefold())
    return (1, 0, path.name.casefold())


def _date_text_from_token(value: str) -> str:
    token = re.sub(r"[^0-9]", "", value)
    if len(token) != 8:
        return ""
    return f"{token[0:4]}-{token[4:6]}-{token[6:8]}"


def _find_latest_size_analysis_workbook(root: Path, profile_id: str) -> Path | None:
    matches: list[tuple[tuple[int, int, str], str, Path]] = []
    seen: set[Path] = set()
    for directory in _size_analysis_dirs(root):
        if not directory.exists():
            continue
        for profile in _candidate_profile_ids(profile_id):
            for path in directory.glob(f"{profile}_*.xlsx"):
                resolved = path.resolve()
                if resolved in seen:
                    continue
                seen.add(resolved)
                matches.append((_size_analysis_sort_key(path), path.name.casefold(), resolved))
    if not matches:
        return None
    return sorted(matches, key=lambda item: (item[0], item[1]))[-1][2]


def _profile_common_workbook_candidates(
    *,
    root: Path,
    profile_id: str,
    brand_label: str,
    workbook_name: str,
    export_size_workbook_name: str,
) -> list[Path]:
    candidates: list[Path] = []
    for base in (root / "Cars_IDCevo", root / "repositories" / "trunk" / "Cars_IDCevo"):
        for profile in _candidate_profile_ids(profile_id):
            common = base / brand_label / profile / "_Common"
            candidates.append(common / workbook_name)
            candidates.append(common / export_size_workbook_name)
    return candidates


def resolve_delivery_checklist_workbook(
    *,
    workspace: Path | str | None = None,
    workbook_path: Path | str | None = None,
    brand: str | None = "BMW",
    profile_id: str = "",
) -> Path:
    if workbook_path is not None:
        return Path(workbook_path).resolve()
    root = _workspace_root(workspace)
    workbook_name = _workbook_name_for_brand(brand)
    export_size_workbook_name = _export_size_workbook_name_for_brand(brand)
    brand_label = _brand_label(brand)
    candidates = [
        root / "repositories" / "trunk" / "Cars" / brand_label / export_size_workbook_name,
        root / "Cars" / brand_label / export_size_workbook_name,
    ]
    latest_size_analysis = _find_latest_size_analysis_workbook(root, profile_id)
    if latest_size_analysis is not None:
        candidates.append(latest_size_analysis)
    candidates.extend(
        [
            root / "repositories" / "trunk" / ".pdx" / "checkers" / "deliveryChecklist" / workbook_name,
            root / ".pdx" / "checkers" / "deliveryChecklist" / workbook_name,
        ]
    )
    candidates.extend(
        _profile_common_workbook_candidates(
            root=root,
            profile_id=profile_id,
            brand_label=brand_label,
            workbook_name=workbook_name,
            export_size_workbook_name=export_size_workbook_name,
        )
    )
    candidates.extend(
        [
            root / export_size_workbook_name,
            root / workbook_name,
        ]
    )
    for candidate in candidates:
        if candidate.exists():
            return candidate.resolve()
    for candidate in candidates:
        if candidate.parent.exists():
            return candidate.resolve()
    return candidates[0].resolve()


def _header_key(value: object) -> str:
    normalized = re.sub(r"[^a-z0-9]+", "", str(value or "").casefold())
    return _HEADER_ALIASES.get(normalized, "")


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="minutes")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, float):
        return format(value, ".12g")
    return str(value).strip()


def _profile_token(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.casefold())


def _candidate_profile_tokens(profile_id: str) -> set[str]:
    tokens = {_profile_token(item) for item in _candidate_profile_ids(profile_id)}
    return {token for token in tokens if token}


def _normalize_status(value: object) -> str:
    raw = _cell_text(value)
    if not raw:
        return "pending"
    normalized = re.sub(r"[^a-z0-9]+", " ", raw.casefold()).strip()
    compact = normalized.replace(" ", "")
    if compact in {"na", "n/a", "notapplicable", "notavailable", "skip", "skipped"}:
        return "not_applicable"
    if compact in {"ok", "pass", "passed", "done", "green", "success", "successful", "yes", "true"}:
        return "passed"
    if compact in {"nok", "notok", "fail", "failed", "failure", "error", "red", "false"}:
        return "failed"
    if "block" in normalized:
        return "blocked"
    if compact in {"pending", "open", "todo", "wip", "waiting"}:
        return "pending"
    return normalized.replace(" ", "_") or "unknown"


def _status_text(status: str) -> str:
    return status.replace("_", " ")


def _recorded_status(value: object) -> str:
    raw = _cell_text(value)
    if not raw:
        return "pending"
    normalized = re.sub(r"[^a-z0-9]+", " ", raw.casefold()).strip()
    compact = normalized.replace(" ", "")
    if compact in {"notrun", "notexecuted", "notstarted"}:
        return "not_run"
    if compact in {"notavailable", "unavailable", "na", "notapplicable"}:
        return "not_available"
    return "recorded"


def _workbook_metadata(workbook_path: Path, *, brand: str | None, row_count: int = 0) -> dict[str, Any]:
    stat = workbook_path.stat() if workbook_path.exists() else None
    return {
        "brand": _brand_label(brand),
        "file_size": int(stat.st_size) if stat else 0,
        "modified_at": datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds") if stat else "",
        "row_count": row_count,
    }


def _missing_payload(
    profile_id: str,
    workbook_path: Path,
    status: str,
    summary: str,
    *,
    brand: str | None,
) -> dict[str, Any]:
    return {
        "profile_id": profile_id.strip(),
        "matched_profile_id": "",
        "status": status,
        "data_available": False,
        "workbook_path": str(workbook_path),
        "worksheet": "",
        "row": 0,
        "last_tested": "",
        "svn_revision": "",
        "changelog_revision": "",
        "workbook_metadata": _workbook_metadata(workbook_path, brand=brand),
        "checks": [],
        "summary": summary,
        "note": "Read-only delivery-checklist evidence guidance; not approval or delivery signoff.",
        "is_approval": False,
    }


def _row_mapping(values: tuple[object, ...], header: dict[int, str]) -> dict[str, object]:
    mapped: dict[str, object] = {}
    for index, canonical in header.items():
        if index < len(values):
            mapped[canonical] = values[index]
    return mapped


def _find_header(values: tuple[object, ...]) -> dict[int, str]:
    header: dict[int, str] = {}
    for index, value in enumerate(values):
        key = _header_key(value)
        if key:
            header[index] = key
    has_profile = any(key in _PROFILE_KEYS for key in header.values())
    has_check = any(key in _CHECK_COLUMNS for key in header.values())
    return header if has_profile and has_check else {}


def _find_export_size_header(values: tuple[object, ...]) -> dict[int, str]:
    header: dict[int, str] = {}
    for index, value in enumerate(values):
        key = _header_key(value)
        if key:
            header[index] = key
    has_evidence = any(key in {"ramses_size", "logic_size"} for key in header.values())
    has_context = any(key in {"last_tested", "svn_revision", "changelog_revision"} for key in header.values())
    return header if has_evidence and has_context else {}


def _overview_total_index(values: tuple[object, ...]) -> int:
    normalized = [re.sub(r"[^a-z0-9]+", "", _cell_text(value).casefold()) for value in values]
    if not normalized or normalized[0] != "variant":
        return -1
    for index, value in enumerate(normalized):
        if value == "total":
            return index
    return -1


def _recorded_check(key: str, label: str, value: object) -> dict[str, str]:
    raw_value = _cell_text(value)
    return {
        "key": key,
        "label": label,
        "status": _recorded_status(value),
        "raw_value": raw_value,
    }


def _overview_delivery_payload(
    *,
    profile: str,
    workbook: Path,
    worksheet_title: str,
    rows: list[tuple[object, ...]],
    brand: str | None,
) -> dict[str, Any] | None:
    if len(rows) < 2:
        return None
    workbook_profile = _cell_text(rows[0][0] if rows[0] else "")
    if _profile_token(workbook_profile) not in _candidate_profile_tokens(profile):
        return None
    header_index = -1
    total_index = -1
    for index, row in enumerate(rows):
        total_index = _overview_total_index(row)
        if total_index >= 0:
            header_index = index
            break
    if header_index < 0:
        return None
    variants = [
        row
        for row in rows[header_index + 1 :]
        if _cell_text(row[0] if row else "")
    ]
    totals = [_cell_text(row[total_index] if total_index < len(row) else "") for row in variants]
    total_detail = ", ".join(item for item in totals if item)
    checks = [
        _recorded_check("variant_count", "Size Analysis Variants", str(len(variants))),
    ]
    if total_detail:
        checks.append(_recorded_check("variant_totals", "Variant Totals", total_detail))
    workbook_date = _date_text_from_token(_filename_date_token(workbook))
    date_text = f" dated {workbook_date}" if workbook_date else ""
    summary = (
        f"Delivery checklist {profile}: size-analysis workbook{date_text} found with "
        f"{len(variants)} variant rows in the Overview sheet."
    )
    return {
        "profile_id": profile,
        "matched_profile_id": workbook_profile,
        "status": "available",
        "data_available": True,
        "workbook_path": str(workbook),
        "worksheet": worksheet_title,
        "row": header_index + 1,
        "last_tested": workbook_date,
        "svn_revision": "",
        "changelog_revision": "",
        "workbook_metadata": _workbook_metadata(workbook, brand=brand, row_count=len(rows)),
        "checks": checks,
        "summary": summary,
        "note": "Read-only delivery-checklist evidence guidance; not approval or delivery signoff.",
        "is_approval": False,
    }


def _profile_in_overview_title(profile: str, title: str) -> bool:
    token = _profile_token(title)
    return any(token == candidate or token.startswith(candidate) for candidate in _candidate_profile_tokens(profile))


def _versioned_overview_delivery_payload(
    *,
    profile: str,
    workbook: Path,
    worksheet_title: str,
    rows: list[tuple[object, ...]],
    brand: str | None,
) -> dict[str, Any] | None:
    if len(rows) < 3:
        return None
    workbook_profile = _cell_text(rows[0][0] if rows[0] else "")
    if not _profile_in_overview_title(profile, workbook_profile):
        return None
    date_row_index = -1
    total_row_index = -1
    for index, row in enumerate(rows):
        label = re.sub(r"[^a-z0-9]+", "", _cell_text(row[0] if row else "").casefold())
        if label == "date":
            date_row_index = index
        elif label == "total":
            total_row_index = index
    if date_row_index <= 0 or total_row_index < 0:
        return None
    header = rows[date_row_index - 1]
    date_row = rows[date_row_index]
    total_row = rows[total_row_index]
    variant_totals: list[str] = []
    tested_dates: list[str] = []
    for column_index in range(1, max(len(header), len(total_row))):
        variant = _cell_text(header[column_index] if column_index < len(header) else "")
        normalized = re.sub(r"[^a-z0-9]+", "", variant.casefold())
        if not variant or normalized in {"min", "max"}:
            continue
        total = _cell_text(total_row[column_index] if column_index < len(total_row) else "")
        if not total:
            continue
        variant_totals.append(f"{variant}={total}")
        tested = _cell_text(date_row[column_index] if column_index < len(date_row) else "")
        if tested and tested not in tested_dates:
            tested_dates.append(tested)
    if not variant_totals:
        return None
    last_tested = tested_dates[0] if tested_dates else ""
    checks = [
        _recorded_check("variant_count", "Size Analysis Variants", str(len(variant_totals))),
        _recorded_check("variant_totals", "Variant Totals", ", ".join(variant_totals)),
    ]
    date_text = f" dated {last_tested}" if last_tested else ""
    summary = (
        f"Delivery checklist {profile}: size-analysis workbook{date_text} found with "
        f"{len(variant_totals)} variant columns in the Overview sheet."
    )
    return {
        "profile_id": profile,
        "matched_profile_id": workbook_profile,
        "status": "available",
        "data_available": True,
        "workbook_path": str(workbook),
        "worksheet": worksheet_title,
        "row": date_row_index + 1,
        "last_tested": last_tested,
        "svn_revision": "",
        "changelog_revision": "",
        "workbook_metadata": _workbook_metadata(workbook, brand=brand, row_count=len(rows)),
        "checks": checks,
        "summary": summary,
        "note": "Read-only delivery-checklist evidence guidance; not approval or delivery signoff.",
        "is_approval": False,
    }


def _sheet_delivery_payload(
    *,
    profile: str,
    workbook: Path,
    worksheet_title: str,
    row_index: int,
    mapped: dict[str, object],
    brand: str | None,
    row_count: int,
) -> dict[str, Any]:
    checks = [
        _recorded_check("ramses_size", "Ramses Size", mapped.get("ramses_size")),
        _recorded_check("logic_size", "Logic Size", mapped.get("logic_size")),
    ]
    summary_parts = [
        f"{item['label']} {_status_text(str(item['status']))}"
        for item in checks
        if item["raw_value"] or item["status"] != "pending"
    ]
    summary = (
        f"Delivery checklist {profile}: {'; '.join(summary_parts)}."
        if summary_parts
        else f"Delivery checklist {profile}: workbook row found, but no export-size values were recorded."
    )
    comment = _cell_text(mapped.get("comment"))
    if comment:
        summary = f"{summary} Comment: {comment}"
    return {
        "profile_id": profile,
        "matched_profile_id": worksheet_title,
        "status": "available",
        "data_available": True,
        "workbook_path": str(workbook),
        "worksheet": worksheet_title,
        "row": row_index,
        "last_tested": _cell_text(mapped.get("last_tested")),
        "svn_revision": _cell_text(mapped.get("svn_revision")),
        "changelog_revision": _cell_text(mapped.get("changelog_revision")),
        "workbook_metadata": _workbook_metadata(workbook, brand=brand, row_count=row_count),
        "checks": checks,
        "summary": summary,
        "note": "Read-only delivery-checklist evidence guidance; not approval or delivery signoff.",
        "is_approval": False,
    }


def read_delivery_checklist(
    *,
    profile_id: str,
    workspace: Path | str | None = None,
    workbook_path: Path | str | None = None,
    brand: str | None = "BMW",
) -> dict[str, Any]:
    profile = profile_id.strip()
    workbook = resolve_delivery_checklist_workbook(
        workspace=workspace,
        workbook_path=workbook_path,
        brand=brand,
        profile_id=profile,
    )
    if not workbook.exists():
        return _missing_payload(
            profile,
            workbook,
            "unavailable",
            (
                f"delivery-checklist data unavailable: workbook not found for {profile or 'profile'}: {workbook}. "
                "BMW export may be complete, but workbook generation is a CI team operation."
            ),
            brand=brand,
        )

    candidate_tokens = _candidate_profile_tokens(profile)
    try:
        loaded = load_workbook(workbook, read_only=True, data_only=True)
    except Exception as exc:
        return _missing_payload(
            profile,
            workbook,
            "unreadable",
            f"delivery-checklist data unavailable: workbook could not be read: {exc}",
            brand=brand,
        )

    try:
        if "Overview" in loaded.sheetnames:
            overview = loaded["Overview"]
            overview_rows = [tuple(row) for row in overview.iter_rows(values_only=True)]
            overview_payload = _overview_delivery_payload(
                profile=profile,
                workbook=workbook,
                worksheet_title=overview.title,
                rows=overview_rows,
                brand=brand,
            )
            if overview_payload is None:
                overview_payload = _versioned_overview_delivery_payload(
                    profile=profile,
                    workbook=workbook,
                    worksheet_title=overview.title,
                    rows=overview_rows,
                    brand=brand,
                )
            if overview_payload is not None:
                return overview_payload
        workbook_row_count = 0
        for worksheet in loaded.worksheets:
            header: dict[int, str] = {}
            header_has_profile = False
            latest_sheet_row: dict[str, object] | None = None
            latest_sheet_row_index = 0
            sheet_matches_profile = _profile_token(worksheet.title) in candidate_tokens
            for row_index, values in enumerate(worksheet.iter_rows(values_only=True), start=1):
                workbook_row_count += 1
                if not header:
                    header = _find_header(values)
                    if header:
                        header_has_profile = True
                    else:
                        header = _find_export_size_header(values)
                    continue
                if not header:
                    continue
                mapped = _row_mapping(values, header)
                if not header_has_profile and sheet_matches_profile:
                    if any(_cell_text(mapped.get(key)) for key in ("last_tested", "svn_revision", "changelog_revision", "ramses_size", "logic_size", "comment")):
                        latest_sheet_row = mapped
                        latest_sheet_row_index = row_index
                    continue
                if not header_has_profile:
                    continue
                profile_value = next((_cell_text(mapped.get(key)) for key in _PROFILE_KEYS if mapped.get(key)), "")
                if not profile_value:
                    continue
                if _profile_token(profile_value) not in candidate_tokens:
                    continue
                checks = [
                    {
                        "key": key,
                        "label": label,
                        "status": _normalize_status(mapped.get(key)),
                        "raw_value": _cell_text(mapped.get(key)),
                    }
                    for key, label in _CHECK_COLUMNS.items()
                    if key in mapped
                ]
                check_summary = "; ".join(
                    f"{item['label']} {_status_text(str(item['status']))}" for item in checks
                )
                summary = f"Delivery checklist {profile}: {check_summary}." if check_summary else f"Delivery checklist {profile}: no check columns found."
                return {
                    "profile_id": profile,
                    "matched_profile_id": profile_value,
                    "status": "available",
                    "data_available": True,
                    "workbook_path": str(workbook),
                    "worksheet": worksheet.title,
                    "row": row_index,
                    "last_tested": _cell_text(mapped.get("last_tested")),
                    "svn_revision": _cell_text(mapped.get("svn_revision")),
                    "changelog_revision": _cell_text(mapped.get("changelog_revision")),
                    "workbook_metadata": _workbook_metadata(workbook, brand=brand, row_count=workbook_row_count),
                    "checks": checks,
                    "summary": summary,
                    "note": "Read-only delivery-checklist evidence guidance; not approval or delivery signoff.",
                    "is_approval": False,
                }
            if latest_sheet_row is not None:
                return _sheet_delivery_payload(
                    profile=profile,
                    workbook=workbook,
                    worksheet_title=worksheet.title,
                    row_index=latest_sheet_row_index,
                    mapped=latest_sheet_row,
                    brand=brand,
                    row_count=workbook_row_count,
                )
    finally:
        loaded.close()

    return _missing_payload(
        profile,
        workbook,
        "unavailable",
        f"delivery-checklist data unavailable: workbook was found, but no row matched {profile}.",
        brand=brand,
    )


def read_delivery_checklists_for_profiles(
    profile_ids: list[str] | tuple[str, ...],
    *,
    workspace: Path | str | None = None,
    workbook_path: Path | str | None = None,
    brand: str | None = "BMW",
) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    for profile_id in profile_ids:
        payload = read_delivery_checklist(profile_id=profile_id, workspace=workspace, workbook_path=workbook_path, brand=brand)
        items.append(payload)
    return items


def delivery_checklist_digest_items(state: dict[str, Any]) -> list[dict[str, Any]]:
    raw_items = state.get("delivery_checklist", [])
    if isinstance(raw_items, dict):
        raw_items = [raw_items]
    if not isinstance(raw_items, list):
        return []
    items: list[dict[str, Any]] = []
    for raw_item in raw_items:
        if not isinstance(raw_item, dict):
            continue
        profile = str(raw_item.get("profile_id", "")).strip() or "profile"
        detail = str(raw_item.get("summary", "")).strip()
        if not detail:
            check_parts = [
                f"{check.get('label', check.get('key', 'check'))} {_status_text(str(check.get('status', 'pending')))}"
                for check in raw_item.get("checks", [])
                if isinstance(check, dict)
            ]
            detail = "; ".join(check_parts)
        items.append(
            {
                "label": f"Delivery checklist {profile}",
                "status": "prepared" if raw_item.get("data_available") else str(raw_item.get("status", "not_available")),
                "detail": detail,
                "source": "delivery_checklist",
                "path": str(raw_item.get("workbook_path", "")).strip(),
                "note": "Read-only delivery-checklist evidence guidance; not approval or delivery signoff.",
            }
        )
    return items


def render_delivery_checklist_markdown(payload: dict[str, Any]) -> str:
    lines = [
        READ_ONLY_BANNER,
        "",
        f"# Delivery Checklist Evidence - {payload.get('profile_id', 'profile')}",
        "",
        f"- Status: `{payload.get('status', 'unknown')}`",
        f"- Data available: `{str(bool(payload.get('data_available'))).lower()}`",
    ]
    workbook_path = str(payload.get("workbook_path", "")).strip()
    if workbook_path:
        lines.append(f"- Workbook: `{workbook_path}`")
    metadata = payload.get("workbook_metadata", {})
    if isinstance(metadata, dict):
        brand = str(metadata.get("brand", "")).strip()
        modified_at = str(metadata.get("modified_at", "")).strip()
        row_count = int(metadata.get("row_count", 0) or 0)
        if brand:
            lines.append(f"- Brand: `{brand}`")
        if modified_at:
            lines.append(f"- Workbook modified: `{modified_at}`")
        if row_count:
            lines.append(f"- Rows scanned: `{row_count}`")
    summary = str(payload.get("summary", "")).strip()
    if summary:
        lines.extend(["", summary])
    checks = payload.get("checks", [])
    if isinstance(checks, list) and checks:
        lines.extend(["", "## Recorded Tests"])
        for check in checks:
            if not isinstance(check, dict):
                continue
            label = str(check.get("label", check.get("key", "check"))).strip()
            status = _status_text(str(check.get("status", "pending")))
            raw_value = str(check.get("raw_value", "")).strip()
            suffix = f" (raw: `{raw_value}`)" if raw_value and raw_value.casefold() != status.casefold() else ""
            lines.append(f"- {label}: `{status}`{suffix}")
    lines.extend(["", "Manual delivery review remains required."])
    return "\n".join(lines).rstrip() + "\n"


def render_delivery_checklist_text(payload: dict[str, Any]) -> str:
    lines = [
        READ_ONLY_BANNER,
        str(payload.get("summary", "Delivery checklist status unavailable.")),
        "Manual delivery review remains required.",
    ]
    workbook_path = str(payload.get("workbook_path", "")).strip()
    if workbook_path:
        lines.append(f"Workbook: {workbook_path}")
    for check in payload.get("checks", []):
        if not isinstance(check, dict):
            continue
        label = str(check.get("label", check.get("key", "check"))).strip()
        status = _status_text(str(check.get("status", "pending")))
        raw_value = str(check.get("raw_value", "")).strip()
        suffix = f" (raw: {raw_value})" if raw_value and raw_value.casefold() != status.casefold() else ""
        lines.append(f"- {label}: {status}{suffix}")
    return "\n".join(lines)
